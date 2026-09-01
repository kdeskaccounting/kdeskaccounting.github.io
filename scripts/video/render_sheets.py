#!/usr/bin/env python3
"""
Render ranges of a LibreOffice-recalculated workbook as Excel-look PNG frames (2400x1350)
using the Playwright headless Chromium shell (cached under ~/Library/Caches/ms-playwright).
Also renders title / outro cards. Every visual element (captions, highlight rings, tab strip)
is drawn here because this ffmpeg build has no text filters.
"""
import openpyxl, html, subprocess, pathlib, datetime, json, re, glob, os
from openpyxl.utils import get_column_letter, range_boundaries, coordinate_to_tuple

FONT_DIR = "/Applications/LibreOffice.app/Contents/Resources/fonts/truetype"
W, H = 2400, 1350
TOP_H, TAB_H, CAP_H = 112, 62, 118          # window chrome, sheet tab strip, caption bar
NAVY, BLUE, YELLOW = "#1F3864", "#2E75B6", "#FFF2CC"

def headless_shell():
    c = sorted(glob.glob(os.path.expanduser(
        "~/Library/Caches/ms-playwright/chromium_headless_shell-*/chrome-headless-shell-mac-arm64/chrome-headless-shell")))
    return c[-1] if c else None

def screenshot(html_path, out_png, w=W, h=H):
    hs = headless_shell()
    exe = hs or "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
    cmd = [exe, "--headless" if hs else "--headless=new", "--disable-gpu", "--hide-scrollbars",
           "--allow-file-access-from-files", "--no-first-run", f"--window-size={w},{h}",
           f"--screenshot={out_png}", f"file://{pathlib.Path(html_path).resolve()}"]
    subprocess.run(cmd, check=True, capture_output=True, timeout=120)

BASE_CSS = f"""
@font-face{{font-family:'Carlito';src:url('file://{FONT_DIR}/Carlito-Regular.ttf')}}
@font-face{{font-family:'Carlito';font-weight:bold;src:url('file://{FONT_DIR}/Carlito-Bold.ttf')}}
@font-face{{font-family:'Carlito';font-style:italic;src:url('file://{FONT_DIR}/Carlito-Italic.ttf')}}
*{{box-sizing:border-box}}
html,body{{margin:0;width:{W}px;height:{H}px;overflow:hidden;background:#0d1a33;font-family:-apple-system,'Helvetica Neue',Arial,sans-serif;color:#1a1a1a}}
.brand{{display:inline-flex;align-items:center;gap:12px;font-weight:700;letter-spacing:.2px}}
.brand i{{display:inline-block;width:22px;height:22px;border-radius:5px;background:linear-gradient(135deg,{BLUE},{NAVY})}}
"""

def rgb(color):
    try:
        v = color.rgb
        if isinstance(v, str) and len(v) >= 6:
            return "#" + v[-6:]
    except Exception:
        pass
    return None

def fmt(v, nf):
    if v is None: return ""
    if isinstance(v, bool): return "TRUE" if v else "FALSE"
    if isinstance(v, (datetime.datetime, datetime.date)):
        n = (nf or "").lower()
        if "mmm" in n and not re.search(r"(?<!m)d", n.replace("dd", "d")): return v.strftime("%b %Y")
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        n = nf or "General"
        if "%" in n:
            m = re.search(r"0\.(0+)%", n); dec = len(m.group(1)) if m else 0
            return f"{v*100:.{dec}f}%"
        if n == "General" or n == "@":
            return str(int(v)) if float(v).is_integer() else f"{v:.2f}"
        m = re.search(r"0\.(0+)", n); dec = len(m.group(1)) if m else 0
        s = f"{abs(v):,.{dec}f}" if ("," in n or "#" in n) else f"{abs(v):.{dec}f}"
        if "$" in n: s = "$" + s
        if v < 0: s = f"({s})" if "(" in n else "-" + s
        return s
    return str(v)

def tab_strip(wb, active):
    tabs = []
    for name in wb.sheetnames:
        if name.startswith("_"): continue
        ws = wb[name]
        tc = None
        try:
            tc = rgb(ws.sheet_properties.tabColor) if ws.sheet_properties.tabColor else None
        except Exception: pass
        cls = "tab active" if name == active else "tab"
        bar = f"border-bottom:5px solid {tc}" if tc else ""
        tabs.append(f'<div class="{cls}" style="{bar}">{html.escape(name)}</div>')
    return "".join(tabs)

def render_sheet(wbv, wbf, sheet, rng, out_png, highlight=(), zoom=1.0, caption="",
                 workbook_name="Workbook.xlsx", html_dir=None, show_caption=True):
    cap_h = CAP_H if show_caption else 0
    ws, wsf = wbv[sheet], wbf[sheet]
    c1, r1, c2, r2 = range_boundaries(rng)
    # merged ranges intersecting the window
    merged, covered = {}, set()
    for m in ws.merged_cells.ranges:
        mc1, mr1, mc2, mr2 = m.bounds
        if mr2 < r1 or mr1 > r2 or mc2 < c1 or mc1 > c2: continue
        a_r, a_c = max(mr1, r1), max(mc1, c1)
        merged[(a_r, a_c)] = (min(mr2, r2) - a_r + 1, min(mc2, c2) - a_c + 1)
        for rr in range(a_r, min(mr2, r2) + 1):
            for cc in range(a_c, min(mc2, c2) + 1):
                if (rr, cc) != (a_r, a_c): covered.add((rr, cc))
    colw = []
    for c in range(c1, c2 + 1):
        L = get_column_letter(c); d = ws.column_dimensions[L] if L in ws.column_dimensions else None
        wch = d.width if (d is not None and d.width) else 8.43
        if d is not None and d.hidden: wch = 0
        colw.append(int(wch * 7.2 + 5))
    rowh = []
    for r in range(r1, r2 + 1):
        d = ws.row_dimensions[r] if r in ws.row_dimensions else None
        hpt = d.height if (d is not None and d.height) else 15.0
        if d is not None and d.hidden: hpt = 0
        px = int(hpt * 1.3333) + 2
        if hpt:  # wrapped text grows the row like Excel's auto-fit would
            for c, cw in zip(range(c1, c2 + 1), colw):
                cell = ws.cell(r, c)
                if cell.alignment is not None and cell.alignment.wrap_text and isinstance(cell.value, str) and cell.value:
                    span = merged.get((r, c)); width = sum(colw[c - c1:c - c1 + (span[1] if span else 1)]) - 10
                    fs = float(cell.font.size or 11) if cell.font is not None else 11.0
                    chars_per_line = max(1, int(width / (fs * 0.62)))
                    lines = sum(max(1, -(-len(part) // chars_per_line)) for part in cell.value.split("\n"))
                    px = max(px, int(lines * fs * 1.45) + 6)
        rowh.append(px)
    gutter, header_h = 48, 26
    table_w, table_h = gutter + sum(colw), header_h + sum(rowh)
    hl = set(highlight)
    hdr = "".join(f'<th style="width:{cw}px;min-width:{cw}px;max-width:{cw}px">{get_column_letter(c)}</th>'
                  for c, cw in zip(range(c1, c2 + 1), colw))
    rows = [f'<tr class="hdr"><th class="gut"></th>{hdr}</tr>']
    for r, rh in zip(range(r1, r2 + 1), rowh):
        tds = [f'<th class="gut" style="height:{rh}px">{r}</th>']
        for c, cw in zip(range(c1, c2 + 1), colw):
            if (r, c) in covered: continue
            cell = ws.cell(r, c); v = fmt(cell.value, cell.number_format)
            st = []
            if cell.fill is not None and cell.fill.fill_type == "solid":
                bg = rgb(cell.fill.fgColor)
                if bg and bg.lower() != "#000000": st.append(f"background:{bg}")
            f = cell.font
            if f is not None:
                if f.bold: st.append("font-weight:bold")
                if f.italic: st.append("font-style:italic")
                if f.size: st.append(f"font-size:{float(f.size)*1.333:.1f}px")
                col = rgb(f.color) if f.color is not None else None
                if col and col.lower() != "#000000": st.append(f"color:{col}")
            al = cell.alignment.horizontal if cell.alignment is not None else None
            isnum = isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)
            if al in ("center", "centerContinuous"): st.append("text-align:center")
            elif al == "right" or (al in (None, "general") and isnum): st.append("text-align:right")
            if cell.alignment is not None and cell.alignment.wrap_text: st.append("white-space:normal")
            if cell.alignment is not None and cell.alignment.indent: st.append(f"padding-left:{6+int(cell.alignment.indent)*10}px")
            span = merged.get((r, c)); attrs = f' rowspan="{span[0]}" colspan="{span[1]}"' if span else ""
            addr = f"{get_column_letter(c)}{r}"
            cls = "c hl" if addr in hl else "c"
            tds.append(f'<td class="{cls}"{attrs} style="{";".join(st)}"><div class="v">{html.escape(v)}</div></td>')
        rows.append("<tr>" + "".join(tds) + "</tr>")
    # fit + focus
    frame_w, frame_h = W, H - TOP_H - TAB_H - cap_h
    s = min((frame_w - 24) / (table_w * 1.012), (frame_h - 24) / (table_h * 1.012)) * zoom
    s = max(0.35, min(s, 3.2))
    if highlight:
        pts = [coordinate_to_tuple(a) for a in highlight]
        fx_t = sum(gutter + sum(colw[:c - c1]) + colw[c - c1] / 2 for _, c in pts) / len(pts)
        fy_t = sum(header_h + sum(rowh[:r - r1]) + rowh[r - r1] / 2 for r, _ in pts) / len(pts)
    else:
        fx_t, fy_t = table_w / 2, table_h / 2
    def place(frame, size, focus, pad):
        if size * s <= frame:
            if pad is None: return (frame - size * s) / 2
            return pad if (frame - size * s) > 2 * pad else (frame - size * s) / 2
        t = frame / 2 - focus * s
        return min(0.0, max(frame - size * s, t))
    tx, ty = place(frame_w, table_w, fx_t, None), place(frame_h, table_h, fy_t, 36.0)
    fx = (tx + fx_t * s) / W; fy = (TOP_H + ty + fy_t * s) / H
    first = highlight[0] if highlight else f"{get_column_letter(c1)}{r1}"
    fv = wsf[first].value
    formula = fv if isinstance(fv, str) else fmt(wsv_cell := ws[first].value, ws[first].number_format)
    if hasattr(fv, "text"): formula = "=" + str(getattr(fv, "text", ""))  # ArrayFormula
    doc = f"""<!doctype html><html><head><meta charset="utf-8"><style>{BASE_CSS}
.top{{height:{TOP_H}px;background:#f3f3f3;border-bottom:1px solid #c8c8c8}}
.title{{height:54px;display:flex;align-items:center;padding:0 28px;font-size:26px;color:#333;background:#e9edf4;gap:18px}}
.title .brand{{font-size:24px;color:{NAVY}}} .title .file{{color:#555}}
.fbar{{height:58px;display:flex;align-items:center;gap:14px;padding:0 20px;font-family:Carlito,Arial;font-size:26px}}
.namebox{{width:190px;height:40px;border:1px solid #bbb;background:#fff;display:flex;align-items:center;padding:0 12px;color:#222}}
.fx{{color:#666;font-style:italic;font-size:24px}} .formula{{flex:1;height:40px;border:1px solid #bbb;background:#fff;display:flex;align-items:center;padding:0 12px;white-space:nowrap;overflow:hidden;color:#222}}
.sheet{{position:relative;width:{frame_w}px;height:{frame_h}px;overflow:hidden;background:#fff}}
table{{position:absolute;left:0;top:0;width:{table_w}px;transform:translate({tx:.2f}px,{ty:.2f}px) scale({s:.5f});transform-origin:0 0;border-collapse:separate;border-spacing:0;table-layout:fixed;font-family:Carlito,Arial,sans-serif;font-size:14.7px;color:#1a1a1a}}
th{{background:#f0f0f0;color:#555;font-weight:400;font-size:13px;border-right:1px solid #cfcfcf;border-bottom:1px solid #cfcfcf;height:{header_h}px;padding:0}}
th.gut{{width:{gutter}px;min-width:{gutter}px;text-align:center;border-right:1px solid #b9b9b9}}
td.c{{border-right:1px solid #dcdcdc;border-bottom:1px solid #dcdcdc;padding:1px 5px;vertical-align:middle;white-space:nowrap;overflow:hidden;text-overflow:clip}}
td.c .v{{overflow:visible;white-space:inherit}}
td.hl{{outline:5px solid #E67E22;outline-offset:-3px;position:relative;z-index:2}}
.tabs{{height:{TAB_H}px;background:#f3f3f3;border-top:1px solid #c8c8c8;display:flex;align-items:flex-end;gap:4px;padding:0 20px}}
.tab{{height:46px;padding:0 22px;display:flex;align-items:center;font-size:23px;color:#444;background:#e4e4e4;border:1px solid #cfcfcf;border-bottom:none;border-radius:6px 6px 0 0}}
.tab.active{{background:#fff;color:{NAVY};font-weight:700;height:52px}}
.cap{{height:{cap_h}px;background:{NAVY};display:flex;align-items:center;justify-content:space-between;padding:0 44px;border-top:8px solid {BLUE}}}
.cap .t{{color:#fff;font-family:Carlito,Arial;font-weight:700;font-size:46px;letter-spacing:.2px}}
.cap .brand{{color:#cfe0f5;font-size:26px;font-weight:600}}
</style></head><body>
<div class="top"><div class="title"><span class="brand"><i></i>KDesk Accounting</span><span class="file">{html.escape(workbook_name)} — {html.escape(sheet)}</span></div>
<div class="fbar"><div class="namebox">{html.escape(first)}</div><span class="fx">fx</span><div class="formula">{html.escape(str(formula))}</div></div></div>
<div class="sheet"><table>{"".join(rows)}</table></div>
<div class="tabs">{tab_strip(wbv, sheet)}</div>
{("<div class=\"cap\"><div class=\"t\">"+html.escape(caption)+"</div><div class=\"brand\"><i></i>kdeskaccounting.com</div></div>") if show_caption else ""}
</body></html>"""
    hp = pathlib.Path(html_dir or pathlib.Path(out_png).parent) / (pathlib.Path(out_png).stem + ".html")
    hp.write_text(doc); screenshot(hp, out_png)
    return {"fx": round(fx, 4), "fy": round(fy, 4), "static": False}

def render_card(kind, out_png, heading, sub="", lines=(), price="", url="", badge="", html_dir=None):
    bullets = "".join(f"<li>{html.escape(l)}</li>" for l in lines)
    body = (f'<div class="badge">{html.escape(badge)}</div>' if badge else "") + \
           f'<h1>{html.escape(heading)}</h1>' + (f'<p class="sub">{html.escape(sub)}</p>' if sub else "") + \
           (f'<ul>{bullets}</ul>' if lines else "") + \
           (f'<div class="price">{html.escape(price)}</div>' if price else "") + \
           (f'<div class="url">{html.escape(url)}</div>' if url else "")
    doc = f"""<!doctype html><html><head><meta charset="utf-8"><style>{BASE_CSS}
body{{background:radial-gradient(1400px 900px at 20% 10%,#2E75B6 0%,{NAVY} 45%,#0d1a33 100%);color:#fff;font-family:Carlito,Arial,sans-serif}}
.wrap{{position:absolute;left:150px;top:0;width:2100px;height:{H}px;display:flex;flex-direction:column;justify-content:center}}
.brand{{position:absolute;left:150px;top:80px;font-size:38px;color:#dbe7f7}} .brand i{{width:34px;height:34px;border-radius:8px;background:#fff}}
.badge{{display:inline-block;align-self:flex-start;background:rgba(255,255,255,.14);border:2px solid rgba(255,255,255,.35);border-radius:999px;padding:12px 30px;font-size:34px;letter-spacing:1px;margin-bottom:34px}}
h1{{font-size:118px;line-height:1.02;margin:0 0 30px;letter-spacing:-1px;max-width:1900px}}
.sub{{font-size:52px;line-height:1.3;margin:0 0 30px;color:#e3ecf9;max-width:1800px}}
ul{{font-size:46px;line-height:1.55;margin:10px 0 30px;padding-left:46px;color:#f0f5fc}}
.price{{font-size:74px;font-weight:700;margin-top:10px}} .price span{{font-size:40px;font-weight:400;color:#cfe0f5}}
.url{{font-size:48px;color:#ffd966;margin-top:18px}}
.foot{{position:absolute;left:150px;bottom:70px;font-size:34px;color:#b9cbe6}}
</style></head><body><div class="brand"><i></i>KDesk Accounting</div><div class="wrap">{body}</div>
<div class="foot">Pure Excel · No macros · Works on Windows and Mac</div></body></html>"""
    hp = pathlib.Path(html_dir or pathlib.Path(out_png).parent) / (pathlib.Path(out_png).stem + ".html")
    hp.write_text(doc); screenshot(hp, out_png)
    return {"fx": 0.5, "fy": 0.5, "static": True}
