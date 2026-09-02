#!/usr/bin/env python3
"""
Vertical YouTube Short (1080x1920, 30 fps, <= 59 s) composed from the existing walkthrough assets:
scene PNGs + narration WAVs + the `short:` block in marketing/video/<slug>/scenes.yaml.
  scripts/video/.venv-tts/bin/python scripts/video/make_short.py --slug asc842
Output: scripts/video/build/<slug>/<slug>-short.mp4 (+ short-review/*.png sample frames).
All text is rendered into PNGs via HTML (this ffmpeg has no drawtext).
"""
import argparse, html, json, math, pathlib, subprocess, sys
import yaml
from PIL import Image, ImageChops
HERE = pathlib.Path(__file__).resolve().parent; REPO = HERE.parents[1]
sys.path.insert(0, str(HERE)); import render_sheets as R
FPS = 30; OUT_W, OUT_H = 1080, 1920; RW, RH = 1296, 2304      # render at 1.2x so zoompan never upsamples
TOP, BOT = 360, 312                                            # bands at render scale (300 / 260 at 1080 wide)
CAP_BAR = 118                                                  # caption bar height on the 2400x1350 scene PNGs
NAVY, BLUE = "#1F3864", "#2E75B6"
GRAD = f"radial-gradient(1100px 700px at 20% 10%, {BLUE} 0%, {NAVY} 45%, #0d1a33 100%)"

def run(cmd):
    r = subprocess.run(cmd, capture_output=True, text=True)
    if r.returncode != 0: raise SystemExit(f"ffmpeg failed:\n{' '.join(cmd)}\n{r.stderr[-1500:]}")
def dur_of(p):
    return float(subprocess.run(["ffprobe", "-v", "error", "-show_entries", "format=duration", "-of", "csv=p=0", str(p)], capture_output=True, text=True).stdout.strip() or 0)

def highlight_bbox(im):
    """Bounding box of the orange (#E67E22) highlight rings drawn by render_sheets, or None."""
    r, g, b = im.split(); tol = 28
    mask = r.point(lambda v: 255 if abs(v - 0xE6) < tol else 0)
    mask = ImageChops.multiply(mask, g.point(lambda v: 255 if abs(v - 0x7E) < tol else 0))
    mask = ImageChops.multiply(mask, b.point(lambda v: 255 if abs(v - 0x22) < tol else 0))
    return mask.getbbox()

def scene_html(hook, caption, img_path, fx, fy, mode="cover", pan=None):
    fit = "object-fit:cover" if mode == "cover" else "object-fit:contain;padding:28px;box-sizing:border-box"
    midbg = "#fff" if mode == "cover" else "#e8eef6"
    if pan:  # fixed magnification: image width pan['w'], positioned so the highlight sits mid-band
        fit = f"position:absolute;left:{pan['left']:.0f}px;top:{pan['top']:.0f}px;width:{pan['w']:.0f}px;height:auto;object-fit:unset"; midbg = "#fff"
    return f"""<!doctype html><html><head><meta charset="utf-8"><style>{R.BASE_CSS}
html,body{{width:{RW}px;height:{RH}px;background:{NAVY}}}
.top{{position:absolute;left:0;top:0;width:{RW}px;height:{TOP}px;background:{GRAD};display:flex;flex-direction:column;justify-content:center;padding:0 64px}}
.brand{{color:#dbe7f7;font-size:30px;margin-bottom:18px}} .brand i{{width:24px;height:24px;background:#fff}}
.hook{{color:#fff;font-family:Carlito,Arial,sans-serif;font-weight:700;font-size:86px;line-height:1.08;letter-spacing:-.5px;display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;overflow:hidden}}
.mid{{position:absolute;left:0;top:{TOP}px;width:{RW}px;height:{RH-TOP-BOT}px;overflow:hidden;background:{midbg}}}
.mid img{{{"width:100%;height:100%;" if not pan else ""}{fit};object-position:{fx*100:.1f}% {fy*100:.1f}%}}
.bot{{position:absolute;left:0;bottom:0;width:{RW}px;height:{BOT}px;background:{NAVY};border-top:10px solid {BLUE};display:flex;flex-direction:column;justify-content:center;padding:0 64px}}
.cap{{color:#fff;font-family:Carlito,Arial,sans-serif;font-weight:700;font-size:58px;line-height:1.15;display:-webkit-box;-webkit-line-clamp:3;-webkit-box-orient:vertical;overflow:hidden}}
.url{{color:#a9bcd8;font-size:30px;margin-top:16px}}
</style></head><body>
<div class="top"><div class="brand"><i></i>KDesk Accounting</div><div class="hook">{html.escape(hook)}</div></div>
<div class="mid"><img src="file://{img_path}"></div>
<div class="bot"><div class="cap">{html.escape(caption)}</div><div class="url">kdeskaccounting.com</div></div>
</body></html>"""

def end_html(cta):
    head, _, link = cta.partition("→"); head = head.strip() or cta; link = link.strip()
    link_html = f'<div class="link">{html.escape(link)}</div>' if link else ""
    return f"""<!doctype html><html><head><meta charset="utf-8"><style>{R.BASE_CSS}
html,body{{width:{RW}px;height:{RH}px;background:{GRAD}}}
.wrap{{position:absolute;left:0;top:0;width:{RW}px;height:{RH}px;display:flex;flex-direction:column;justify-content:center;align-items:center;text-align:center;padding:0 90px}}
.brand{{color:#dbe7f7;font-size:40px;margin-bottom:60px}} .brand i{{width:32px;height:32px;background:#fff}}
.cta{{color:#fff;font-family:Carlito,Arial,sans-serif;font-weight:700;font-size:92px;line-height:1.12}}
.link{{color:#ffd966;font-family:Carlito,Arial,sans-serif;font-weight:700;font-size:58px;line-height:1.2;margin-top:44px;word-break:break-all}}
.sub{{color:#dbe7f7;font-size:40px;margin-top:56px}}
</style></head><body><div class="wrap"><div class="brand"><i></i>KDesk Accounting</div><div class="cta">{html.escape(head)}</div>{link_html}<div class="sub">Pure Excel · No macros · Windows &amp; Mac</div></div></body></html>"""

def main():
    ap = argparse.ArgumentParser(); ap.add_argument("--slug", required=True); ap.add_argument("--crf", type=int, default=26); a = ap.parse_args()
    slug = a.slug; spec = yaml.safe_load(open(REPO / "marketing/video" / slug / "scenes.yaml")); sh = spec["short"]
    build = HERE / "build" / slug; work = build / "short"; work.mkdir(parents=True, exist_ok=True)
    focus = json.load(open(build / "frames/focus.json")); durs = json.load(open(build / "audio/durations.json"))
    parts = []
    ranges = {str(k): v for k, v in (sh.get("ranges") or {}).items()}
    wbv = wbf = None
    for k, idx in enumerate(sh["scenes"]):
        sc = spec["scenes"][idx]; mode = "cover"; fx = fy = 0.5; pan = None
        if str(idx) in ranges:  # dedicated portrait-friendly render of a narrower range, trimmed to the table
            if wbv is None:
                from openpyxl import load_workbook
                rec = build / "recalc" / "src.xlsx"; wbv = load_workbook(rec, data_only=True); wbf = load_workbook(rec)
            raw = work / f"portrait_{idx:02d}.png"
            foc = R.render_sheet(wbv, wbf, sc["sheet"], ranges[str(idx)], raw, tuple(sc.get("highlight", [])), 1.0, "", spec.get("workbook_name", "Workbook.xlsx"), show_caption=False)
            px, py = foc["fx"] * R.W, foc["fy"] * R.H
            im = Image.open(raw).convert("RGB").crop((0, R.TOP_H, R.W, R.H - R.TAB_H)); ox, oy = 0, R.TOP_H
            bbox = ImageChops.difference(im, Image.new("RGB", im.size, (255, 255, 255))).getbbox()
            if bbox:
                x0, y0 = max(0, bbox[0] - 20), max(0, bbox[1] - 20); im = im.crop((x0, y0, min(im.width, bbox[2] + 48), min(im.height, bbox[3] + 20))); ox += x0; oy += y0
            cropped = work / f"crop_{idx:02d}.png"; im.save(cropped); mode = "pan"
            fx = (px - ox) / im.width; fy = (py - oy) / im.height
            band_w, band_h = RW, RH - TOP - BOT
            # magnify as far as 1.85x the band width, but never so far that the orange highlight ring(s) leave the band
            hb = highlight_bbox(im)
            scale = 1.85 * RW / im.width
            if hb:
                hw, hh = hb[2] - hb[0], hb[3] - hb[1]
                scale = min(scale, (band_w - 96) / max(hw, 1), (band_h - 96) / max(hh, 1)); scale = max(scale, min(1.0, RW / im.width))
                fx = ((hb[0] + hb[2]) / 2) / im.width; fy = ((hb[1] + hb[3]) / 2) / im.height
            w_img = im.width * scale; h_img = im.height * scale
            left = (band_w - w_img) / 2 if w_img <= band_w else min(0.0, max(band_w - w_img, band_w / 2 - fx * w_img))
            top = (band_h - h_img) / 2 if h_img <= band_h else min(0.0, max(band_h - h_img, band_h / 2 - fy * h_img))
            pan = {"w": w_img, "left": left, "top": top}
        else:
            src = Image.open(build / "frames" / f"scene_{idx:02d}.png"); crop = src.crop((0, 0, src.width, src.height - CAP_BAR))
            cropped = work / f"crop_{idx:02d}.png"; crop.save(cropped)
            f = focus.get(str(idx), {}); fx = min(0.72, max(0.30, float(f.get("fx", 0.5)))); fy = min(0.60, max(0.20, float(f.get("fy", 0.5)) * src.height / crop.height))
        hp = work / f"scene_{k}.html"; hp.write_text(scene_html(sh["hook"], sc.get("caption", ""), cropped.resolve(), fx, fy, mode, pan))
        png = work / f"scene_{k}.png"; R.screenshot(hp, png, RW, RH)
        wav = build / "audio" / f"scene_{idx:02d}.wav"; adur = float(durs.get(str(idx), 0) or dur_of(wav)); dur = adur + 0.6; n = math.ceil(dur * FPS)
        zmax = 1.06; dz = (zmax - 1.0) / n
        vf = (f"scale={RW}:{RH}:flags=lanczos,zoompan=z='min(zoom+{dz:.7f},{zmax})':x='iw/2-(iw/zoom/2)':y='ih/2-(ih/zoom/2)':d={n}:s={OUT_W}x{OUT_H}:fps={FPS},"
              f"fade=t=in:st=0:d=0.3,fade=t=out:st={max(0.0, dur-0.3):.3f}:d=0.3,format=yuv420p")
        out = work / f"scene_{k}.mp4"
        run(["ffmpeg", "-y", "-loglevel", "error", "-i", str(png), "-i", str(wav), "-filter_complex",
             f"[0:v]{vf}[v];[1:a]apad=pad_dur=2,afade=t=in:d=0.05,aformat=sample_rates=48000:channel_layouts=stereo[a]",
             "-map", "[v]", "-map", "[a]", "-t", f"{dur:.3f}", "-c:v", "libx264", "-preset", "medium", "-crf", str(a.crf), "-r", str(FPS), "-c:a", "aac", "-b:a", "128k", str(out)])
        parts.append(out); print(f"scene {idx:02d}: {dur:.1f}s -> {out.name}", flush=True)
    hp = work / "end.html"; hp.write_text(end_html(sh["cta"])); png = work / "end.png"; R.screenshot(hp, png, RW, RH)
    out = work / "end.mp4"; n = int(1.5 * FPS)
    run(["ffmpeg", "-y", "-loglevel", "error", "-i", str(png), "-f", "lavfi", "-i", "anullsrc=r=48000:cl=stereo", "-filter_complex",
         f"[0:v]scale={RW}:{RH},zoompan=z='1':d={n}:s={OUT_W}x{OUT_H}:fps={FPS},fade=t=in:st=0:d=0.3,format=yuv420p[v]", "-map", "[v]", "-map", "1:a", "-t", "1.5",
         "-c:v", "libx264", "-preset", "medium", "-crf", str(a.crf), "-r", str(FPS), "-c:a", "aac", "-b:a", "128k", str(out)])
    parts.append(out)
    lst = work / "concat.txt"; lst.write_text("".join(f"file '{p.resolve()}'\n" for p in parts))
    final = build / f"{slug}-short.mp4"
    run(["ffmpeg", "-y", "-loglevel", "error", "-f", "concat", "-safe", "0", "-i", str(lst), "-c:v", "copy", "-af", "loudnorm=I=-16:TP=-1.5:LRA=11", "-c:a", "aac", "-b:a", "128k", "-movflags", "+faststart", str(final)])
    total = dur_of(final)
    if total > 59.5: raise SystemExit(f"Short too long: {total:.1f}s (>59 s) — pick shorter scenes")
    rev = build / "short-review"; rev.mkdir(exist_ok=True)
    for name, t in (("t01", 1.0), ("mid", total / 2), ("end", max(0.0, total - 1.0))):
        run(["ffmpeg", "-y", "-loglevel", "error", "-ss", f"{t:.2f}", "-i", str(final), "-frames:v", "1", str(rev / f"{name}.png")])
    probe = subprocess.run(["ffprobe", "-v", "error", "-select_streams", "v:0", "-show_entries", "stream=width,height,r_frame_rate", "-show_entries", "format=duration,size", "-of", "default=nw=1", str(final)], capture_output=True, text=True).stdout.replace("\n", " ")
    print("FINAL:", final, "|", probe.strip())
if __name__ == "__main__": main()
