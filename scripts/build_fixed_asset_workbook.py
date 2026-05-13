#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "openpyxl>=3.1",
# ]
# ///
"""
Build the KDesk Accounting Fixed Asset Rollforward Workbook (paid + free).

Source of truth: marketing/products/fixed-asset-rollforward-spec.md (v2).

Produces:
- dist/templates/fixed-asset-rollforward-v1.xlsx   (paid: 50 assets x 120 months)
- dist/templates/fixed-asset-rollforward-free-v1.xlsx (free: 5 assets x 36 months)

Invocation:
    uv run scripts/build_fixed_asset_workbook.py                  # build both
    uv run scripts/build_fixed_asset_workbook.py --variant=paid   # paid only
    uv run scripts/build_fixed_asset_workbook.py --variant=free   # free only

Function library restricted to Excel 2016/365/Mac safe set (no XLOOKUP, LET,
LAMBDA, IFS, OFFSET, INDIRECT, dynamic-array spills). All formulas tested for
syntactic round-trip via openpyxl re-load.

Sheet protection password: kdesk2026
Opening Accum Depr (Register col S) sub-password: kdesk-opening-2026
"""
from __future__ import annotations

import argparse
import datetime as dt
import pathlib
import sys
from typing import Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Protection,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins, PrintOptions

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

REPO_ROOT = pathlib.Path(__file__).resolve().parents[1]
DIST = REPO_ROOT / "dist" / "templates"

SHEET_PW = "kdesk2026"
OPENING_PW = "kdesk-opening-2026"

# Colors
COLOR_HEADER_BG = "FF1F4E78"     # navy
COLOR_HEADER_FG = "FFFFFFFF"
COLOR_SECTION_BG = "FFD9E1F2"    # light blue band
COLOR_EXAMPLE_BG = "FFEEEEEE"    # gray for example rows
COLOR_INPUT_BG = "FFFFF2CC"      # pale yellow for input cells
COLOR_CALC_BG = "FFFFFFFF"
COLOR_GREEN = "FFC6EFCE"
COLOR_YELLOW = "FFFFEB9C"
COLOR_RED = "FFFFC7CE"
COLOR_BORDER = "FFB7B7B7"
COLOR_BRAND = "FF1F4E78"

# Lists
CATEGORIES = [
    "Computer Equipment",
    "Furniture & Fixtures",
    "Leasehold Improvements",
    "Machinery",
    "Vehicles",
    "Other",
]
METHODS_PAID = ["Straight-Line", "DDB", "SYD", "UoP"]
METHODS_FREE = ["Straight-Line", "DDB", "SYD"]
CONVENTIONS = ["Full Month", "Mid-Month"]
DISPOSAL_TYPES = ["Sale", "Retirement", "Trade-in", "Write-off", "Theft-loss"]
EVIDENCE_TYPES = [
    "Bill of Sale",
    "Scrap Receipt",
    "Write-off Memo",
    "Insurance Claim",
    "Police Report",
    "Other",
]
GL_SYSTEMS = ["Generic", "QuickBooks", "NetSuite", "Sage Intacct", "Xero"]

# Default Policy (rows == CATEGORIES)
POLICY = [
    ("Computer Equipment", 36, "Straight-Line", 0.0),
    ("Furniture & Fixtures", 84, "Straight-Line", 0.0),
    ("Leasehold Improvements", 60, "Straight-Line", 0.0),
    ("Machinery", 84, "Straight-Line", 0.0),
    ("Vehicles", 60, "Straight-Line", 0.10),
    ("Other", 60, "Straight-Line", 0.0),
]

# Example asset rows (gray-filled, show all 4 methods in paid; 2 rows in free)
# (AssetID, Description, Category, Dept, Vendor, Invoice#, PO#, Approver,
#  ApprovalDate, AcquisitionDate, InServiceDate, Cost, Salvage, Life,
#  Method, DDB Rate override, Total Production Units, Convention,
#  Opening Accum Depr, Opening Period, LHI Lease End, LHI Renewal Y/N,
#  Revised Useful Life, Revision Effective Period, Tax Memo, Notes)
EXAMPLES_PAID = [
    ("FA-0001", "MacBook Pro M3 14\"",        "Computer Equipment",    "Engineering", "Apple Inc.",         "INV-10142", "PO-2026-001", "S. Michels", dt.date(2026, 1, 5), dt.date(2026, 1, 10), dt.date(2026, 1, 15), 2599.00, 0.0,     36, "Straight-Line", None,   None,  "Full Month", 0.0, dt.date(2026, 1, 1), None, "Y", None, None, "Bonus dep §168(k) on tax side", "Engineering laptop"),
    ("FA-0002", "Conference Room Table",      "Furniture & Fixtures",  "G&A",         "Herman Miller",      "INV-887",   "PO-2026-002", "S. Michels", dt.date(2026, 1, 14), dt.date(2026, 1, 20), dt.date(2026, 2, 1),  4200.00, 0.0,     84, "Straight-Line", None,   None,  "Full Month", 0.0, dt.date(2026, 1, 1), None, "Y", None, None, "",                              "HQ conference room"),
    ("FA-0003", "Office Buildout — 5th floor", "Leasehold Improvements","G&A",         "Pacific Contractors","INV-2201",  "PO-2026-003", "CFO Sign-off", dt.date(2025, 11, 10), dt.date(2025, 12, 1), dt.date(2026, 1, 1),  85000.00, 0.0,     60, "Straight-Line", None,   None,  "Mid-Month", 0.0, dt.date(2026, 1, 1), dt.date(2030, 12, 31), "Y", None, None, "",            "Lease ends 2030-12; LHI life = MIN(60, months-to-lease-end)"),
    ("FA-0004", "Server Rack — Colo Cabinet", "Machinery",             "Engineering", "ServerMonkey",       "INV-50301", "PO-2026-004", "S. Michels", dt.date(2026, 2, 1), dt.date(2026, 2, 5), dt.date(2026, 2, 10), 18500.00, 1500.00, 84, "DDB",           None,   None,  "Full Month", 0.0, dt.date(2026, 1, 1), None, "Y", None, None, "",                                       "Demonstrates DDB method"),
    ("FA-0005", "Delivery Vehicle — Ford Transit", "Vehicles",          "Operations",  "Ford Auto Group",    "INV-VEH-77","PO-2026-005", "COO Sign-off", dt.date(2026, 2, 12), dt.date(2026, 2, 18), dt.date(2026, 3, 1),  42000.00, 4200.00, 60, "SYD",           None,   None,  "Full Month", 0.0, dt.date(2026, 1, 1), None, "Y", None, None, "Sec §179 partial on tax side; book SYD", "SYD method, salvage 10%"),
    ("FA-0006", "CNC Mill — Production Line",  "Machinery",             "Operations",  "Haas Automation",    "INV-CNC-1", "PO-2026-006", "COO Sign-off", dt.date(2026, 3, 5), dt.date(2026, 3, 12), dt.date(2026, 3, 15), 87000.00, 7000.00, 96, "UoP",           None,   100000, "Full Month", 0.0, dt.date(2026, 1, 1), None, "Y", None, None, "",                                       "UoP method; 100k unit lifetime"),
]
EXAMPLES_FREE = [
    EXAMPLES_PAID[0],  # SL Computer
    EXAMPLES_PAID[3],  # DDB Machinery
]

# Disposal Log examples
DISPOSALS_PAID = [
    # (AssetID, DispDate, Type, CommSubstance, Proceeds, AuthName, AuthDate,
    #  EvidenceType, EvidenceRef, Counterparty, Notes)
    ("FA-0005", dt.date(2027, 6, 15), "Sale", "N/A", 32000.00, "COO Sign-off", dt.date(2027, 6, 14), "Bill of Sale", "BoS-2027-08", "Acme Delivery LLC", "Sold mid-life; gain calc"),
]
DISPOSALS_FREE: list = []


# Schedule depth by variant
SCHED_DEPTH = {"paid": 120, "free": 36}
MAX_ASSETS = {"paid": 50, "free": 5}


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

THIN = Side(style="thin", color=COLOR_BORDER)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill(fill_type="solid", start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG)
SECTION_FILL = PatternFill(fill_type="solid", start_color=COLOR_SECTION_BG, end_color=COLOR_SECTION_BG)
EXAMPLE_FILL = PatternFill(fill_type="solid", start_color=COLOR_EXAMPLE_BG, end_color=COLOR_EXAMPLE_BG)
INPUT_FILL = PatternFill(fill_type="solid", start_color=COLOR_INPUT_BG, end_color=COLOR_INPUT_BG)
GREEN_FILL = PatternFill(fill_type="solid", start_color=COLOR_GREEN, end_color=COLOR_GREEN)
YELLOW_FILL = PatternFill(fill_type="solid", start_color=COLOR_YELLOW, end_color=COLOR_YELLOW)
RED_FILL = PatternFill(fill_type="solid", start_color=COLOR_RED, end_color=COLOR_RED)

HEADER_FONT = Font(bold=True, color=COLOR_HEADER_FG, size=11)
TITLE_FONT = Font(bold=True, color=COLOR_BRAND, size=16)
SECTION_FONT = Font(bold=True, color=COLOR_HEADER_BG, size=12)
BODY_FONT = Font(size=10)
SMALL_FONT = Font(size=9, italic=True, color="FF555555")

LOCKED = Protection(locked=True, hidden=False)
UNLOCKED = Protection(locked=False, hidden=False)


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER_ALL


def style_section(cell):
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER_ALL


def style_input(cell):
    cell.fill = INPUT_FILL
    cell.protection = UNLOCKED
    cell.border = BORDER_ALL


def style_calc(cell):
    cell.protection = LOCKED
    cell.border = BORDER_ALL


def set_col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ---------------------------------------------------------------------------
# Tab builders
# ---------------------------------------------------------------------------

def build_readme(wb: Workbook, variant: str):
    ws = wb.create_sheet("README", 0)
    ws.sheet_properties.tabColor = COLOR_BRAND

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110

    ws["B2"] = "KDesk Accounting — Fixed Asset Rollforward Workbook"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = f"Version 1.0  |  {'Free (5 assets x 36 months)' if variant=='free' else 'Paid (50 assets x 120 months)'}  |  US GAAP (ASC 360) — Book Depreciation"
    ws["B3"].font = SMALL_FONT

    row = 5
    ws[f"B{row}"] = "Quickstart (5 steps)"
    style_section(ws[f"B{row}"])
    row += 1
    steps = [
        "1. Open the Setup tab. Enter your company name, fiscal year start, reporting period, and performance materiality. Confirm or override the GL account codes and Policy table.",
        "2. On the Asset Register, the first rows are EXAMPLE assets (gray-filled). Clear them before entering your own — your assets go in the white rows below.",
        "3. Load your assets. Required fields are highlighted yellow. Category drives default useful life / method / salvage from the Policy table.",
        "4. Log any disposals on the Disposal Log tab. Required: Asset ID, Disposal Date, Type, Proceeds, Authorization, Evidence Type/Reference.",
        "5. On Setup, set the Reporting Period to the close period you want. Open the Reconciliation tab — every check should show TIES (green). Then export JE Generator to your GL.",
    ]
    for s in steps:
        ws[f"B{row}"] = s
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1
    ws[f"B{row}"] = "Walkthrough video"
    style_section(ws[f"B{row}"])
    row += 1
    ws[f"B{row}"] = "60-90 sec Loom walkthrough: https://www.loom.com/share/PLACEHOLDER  (URL added post-build)"
    row += 2

    ws[f"B{row}"] = "What this is NOT"
    style_section(ws[f"B{row}"])
    row += 1
    out_of_scope = [
        "* MACRS / tax depreciation books — this is book GAAP only.",
        "* Bonus depreciation (§168(k)) and §179 elections — tax book is out of scope. Free-text memo column provided for cross-reference; this is not tax advice.",
        "* Impairment testing (ASC 360-10-35-17 through -35) — manual write-down via Disposal Log only.",
        "* Componentization (ASC 360-10-35-4) — single asset = single schedule.",
        "* Asset Retirement Obligations (ASC 410-20).",
        "* Lessor accounting.",
        "* Asset transfers between categories — workflow is re-key the asset.",
        "* Asset revaluation upward — GAAP prohibits this for PP&E.",
    ]
    for s in out_of_scope:
        ws[f"B{row}"] = s
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    row += 1
    ws[f"B{row}"] = "Workbook protection"
    style_section(ws[f"B{row}"])
    row += 1
    ws[f"B{row}"] = f"Sheet protection password: {SHEET_PW}"
    row += 1
    ws[f"B{row}"] = f"Opening Accum Depr (Register col S) sub-password: {OPENING_PW}"
    row += 1
    ws[f"B{row}"] = "Protection scope: formula cells are locked across all tabs; input cells are unlocked. Unprotect a sheet via Review > Unprotect Sheet."
    ws[f"B{row}"].alignment = Alignment(wrap_text=True)
    row += 2

    ws[f"B{row}"] = "Support"
    style_section(ws[f"B{row}"])
    row += 1
    ws[f"B{row}"] = "hello@kdeskaccounting.com  |  https://kdeskaccounting.com"
    row += 2

    ws[f"B{row}"] = "Change history"
    style_section(ws[f"B{row}"])
    row += 1
    ws[f"B{row}"] = "v1.0 — 2026-05-12 — Initial release"
    row += 1

    # Lock everything; README has no inputs.
    for r in ws.iter_rows():
        for c in r:
            c.protection = LOCKED
    ws.protection.set_password(SHEET_PW)
    ws.protection.sheet = True


# ---------------------------------------------------------------------------
# Tab 1: Setup
# ---------------------------------------------------------------------------

def build_setup(wb: Workbook, variant: str):
    ws = wb.create_sheet("Setup")
    ws.sheet_properties.tabColor = "FF8FAADC"

    set_col_widths(ws, {
        "A": 28, "B": 22, "C": 22, "D": 22, "E": 22, "F": 22,
    })

    # Company section
    ws["A1"] = "Company & Reporting"
    style_section(ws["A1"])
    ws.merge_cells("A1:F1")

    ws["A2"] = "Company name"
    ws["B2"] = "Your Company, Inc."  # input
    style_input(ws["B2"])

    ws["A3"] = "Fiscal year start (date)"
    ws["B3"] = dt.date(2026, 1, 1)
    ws["B3"].number_format = "yyyy-mm-dd"
    style_input(ws["B3"])

    ws["A4"] = "Reporting Period (Month End)"
    ws["B4"] = dt.date(2026, 12, 31)
    ws["B4"].number_format = "yyyy-mm-dd"
    style_input(ws["B4"])

    ws["A5"] = "Performance Materiality (USD)"
    ws["B5"] = 50000
    ws["B5"].number_format = "#,##0.00"
    style_input(ws["B5"])

    # Conventions
    ws["A7"] = "Conventions"
    style_section(ws["A7"])
    ws.merge_cells("A7:F7")

    ws["A8"] = "Default first-period convention"
    ws["B8"] = "Full Month"
    style_input(ws["B8"])

    ws["A9"] = "Default depreciation method"
    ws["B9"] = "Straight-Line"
    style_input(ws["B9"])

    if variant == "paid":
        ws["A11"] = "GL Dimension Labels (relabels JE Generator columns I/J/K)"
        ws["B11"] = "Generic"
        style_input(ws["B11"])
        ws["C11"] = "Relabels 3 dimension headers only; JE body is generic — map to your GL's import template."
        ws["C11"].font = Font(italic=True, color="FF777777", size=9)
    else:
        ws["A11"] = "GL Dimension Labels"
        ws["B11"] = "Generic"
        ws["B11"].font = Font(italic=True, color="FF777777")
        ws["B11"].protection = LOCKED

    # GL Accounts
    ws["A13"] = "GL Accounts (named ranges)"
    style_section(ws["A13"])
    ws.merge_cells("A13:F13")
    gl_rows = [
        ("Computer Equipment (asset)", "1510", "GL_ComputerEq"),
        ("Furniture & Fixtures (asset)", "1520", "GL_Furniture"),
        ("Leasehold Improvements (asset)", "1530", "GL_LHI"),
        ("Machinery (asset)", "1540", "GL_Machinery"),
        ("Vehicles (asset)", "1550", "GL_Vehicles"),
        ("Other (asset)", "1590", "GL_Other"),
        ("Accumulated Depreciation", "1599", "GL_AccumDepr"),
        ("Depreciation Expense", "6200", "GL_DeprExpense"),
        ("Gain on Disposal", "7300", "GL_DispGain"),
        ("Loss on Disposal", "7310", "GL_DispLoss"),
    ]
    for i, (label, code, _name) in enumerate(gl_rows):
        r = 14 + i
        ws[f"A{r}"] = label
        ws[f"B{r}"] = code
        style_input(ws[f"B{r}"])

    # Policy table
    pol_r0 = 14 + len(gl_rows) + 1
    ws[f"A{pol_r0}"] = "Policy table (defaults pulled into Asset Register)"
    style_section(ws[f"A{pol_r0}"])
    ws.merge_cells(f"A{pol_r0}:F{pol_r0}")

    headers = ["Category", "Default Useful Life (months)", "Default Method", "Default Salvage %"]
    for j, h in enumerate(headers):
        c = ws.cell(row=pol_r0 + 1, column=1 + j, value=h)
        style_header(c)

    for i, (cat, life, method, salv) in enumerate(POLICY):
        r = pol_r0 + 2 + i
        ws.cell(row=r, column=1, value=cat).protection = LOCKED  # category label locked
        ws.cell(row=r, column=2, value=life)
        ws.cell(row=r, column=3, value=method)
        ws.cell(row=r, column=4, value=salv)
        ws.cell(row=r, column=4).number_format = "0.00%"
        for j in range(2, 5):
            style_input(ws.cell(row=r, column=j))
            ws.cell(row=r, column=j).border = BORDER_ALL

    # Closed Periods
    cp_r0 = pol_r0 + 2 + len(POLICY) + 1
    ws[f"A{cp_r0}"] = "Closed Periods"
    style_section(ws[f"A{cp_r0}"])
    ws.merge_cells(f"A{cp_r0}:F{cp_r0}")
    cp_hdr = ["Period (Month End)", "Closed Date", "Closed By"]
    for j, h in enumerate(cp_hdr):
        c = ws.cell(row=cp_r0 + 1, column=1 + j, value=h)
        style_header(c)
    cp_rows_n = 36
    for i in range(cp_rows_n):
        r = cp_r0 + 2 + i
        for j in range(1, 4):
            cell = ws.cell(row=r, column=j)
            style_input(cell)
            if j <= 2:
                cell.number_format = "yyyy-mm-dd"
    # Pre-fill one example
    ws.cell(row=cp_r0 + 2, column=1, value=dt.date(2026, 1, 31)).number_format = "yyyy-mm-dd"
    ws.cell(row=cp_r0 + 2, column=2, value=dt.date(2026, 2, 12)).number_format = "yyyy-mm-dd"
    ws.cell(row=cp_r0 + 2, column=3, value="Stephen Michels")

    # Hidden category/method/convention lists (for data validations) — drop on the right
    list_col = 8  # column H
    ws.cell(row=1, column=list_col, value="CategoryList").font = SMALL_FONT
    for i, c in enumerate(CATEGORIES):
        ws.cell(row=2 + i, column=list_col, value=c).protection = LOCKED
    ws.cell(row=1, column=list_col + 1, value="MethodList").font = SMALL_FONT
    methods = METHODS_PAID if variant == "paid" else METHODS_FREE
    for i, m in enumerate(methods):
        ws.cell(row=2 + i, column=list_col + 1, value=m).protection = LOCKED
    ws.cell(row=1, column=list_col + 2, value="ConventionList").font = SMALL_FONT
    for i, c in enumerate(CONVENTIONS):
        ws.cell(row=2 + i, column=list_col + 2, value=c).protection = LOCKED
    ws.cell(row=1, column=list_col + 3, value="DisposalTypeList").font = SMALL_FONT
    for i, d in enumerate(DISPOSAL_TYPES):
        ws.cell(row=2 + i, column=list_col + 3, value=d).protection = LOCKED
    ws.cell(row=1, column=list_col + 4, value="EvidenceTypeList").font = SMALL_FONT
    for i, e in enumerate(EVIDENCE_TYPES):
        ws.cell(row=2 + i, column=list_col + 4, value=e).protection = LOCKED
    if variant == "paid":
        ws.cell(row=1, column=list_col + 5, value="GLSystemList").font = SMALL_FONT
        for i, s in enumerate(GL_SYSTEMS):
            ws.cell(row=2 + i, column=list_col + 5, value=s).protection = LOCKED
    # Hide helper columns
    for col_letter in ["H", "I", "J", "K", "L", "M"]:
        ws.column_dimensions[col_letter].hidden = True

    # Data validations for the input dropdowns on Setup
    conv_dv = DataValidation(type="list", formula1=f"=ConventionList", allow_blank=True)
    conv_dv.add(ws["B8"])
    ws.add_data_validation(conv_dv)
    meth_dv = DataValidation(type="list", formula1=f"=MethodList", allow_blank=True)
    meth_dv.add(ws["B9"])
    ws.add_data_validation(meth_dv)
    if variant == "paid":
        gls_dv = DataValidation(type="list", formula1=f"=GLSystemList", allow_blank=True)
        gls_dv.add(ws["B11"])
        ws.add_data_validation(gls_dv)
    fy_dv = DataValidation(type="list", formula1='"2026-01-01,2026-04-01,2026-07-01,2026-10-01"', allow_blank=True)
    fy_dv.add(ws["B3"])
    ws.add_data_validation(fy_dv)

    ws.protection.set_password(SHEET_PW)
    ws.protection.sheet = True

    # Remember anchor positions for named ranges
    ws._anchors = {
        "ReportingPeriodEnd": "B4",
        "Co_Name": "B2",
        "FY_Start": "B3",
        "Materiality": "B5",
        "DefaultConvention": "B8",
        "DefaultMethod": "B9",
        "GLSystem": "B11" if variant == "paid" else None,
        "GL_ComputerEq": "B14",
        "GL_Furniture": "B15",
        "GL_LHI": "B16",
        "GL_Machinery": "B17",
        "GL_Vehicles": "B18",
        "GL_Other": "B19",
        "GL_AccumDepr": "B20",
        "GL_DeprExpense": "B21",
        "GL_DispGain": "B22",
        "GL_DispLoss": "B23",
        "Policy_Cat_Start_Row": pol_r0 + 2,
        "Policy_Cat_End_Row": pol_r0 + 2 + len(POLICY) - 1,
        "ClosedPeriods_Start_Row": cp_r0 + 2,
        "ClosedPeriods_End_Row": cp_r0 + 2 + cp_rows_n - 1,
        "ListsCol": list_col,
        "CategoryList_Range": (list_col, 2, list_col, 1 + len(CATEGORIES)),
        "MethodList_Range": (list_col + 1, 2, list_col + 1, 1 + len(methods)),
        "ConventionList_Range": (list_col + 2, 2, list_col + 2, 1 + len(CONVENTIONS)),
        "DisposalTypeList_Range": (list_col + 3, 2, list_col + 3, 1 + len(DISPOSAL_TYPES)),
        "EvidenceTypeList_Range": (list_col + 4, 2, list_col + 4, 1 + len(EVIDENCE_TYPES)),
    }
    if variant == "paid":
        ws._anchors["GLSystemList_Range"] = (list_col + 5, 2, list_col + 5, 1 + len(GL_SYSTEMS))


# ---------------------------------------------------------------------------
# Tab 2: Asset Register
# ---------------------------------------------------------------------------

REGISTER_HEADERS = [
    "Asset ID", "Description", "Category", "Department",                   # A B C D
    "Vendor", "Invoice #", "PO #", "CapEx Approver", "Approval Date",      # E F G H I
    "Acquisition Date", "In-Service Date",                                  # J K
    "Cost", "Salvage Value", "Useful Life (months)", "Method",              # L M N O
    "DDB Rate (override)", "Total Production Units",                        # P Q
    "First Period Convention", "Opening Accum Depr", "Opening Period",      # R S T
    "Depr Start Date",                                                       # U
    "LHI Lease End Date", "LHI Renewal Reasonably Certain?",                # V W
    "Effective Useful Life (calc)",                                          # X
    "Revised Useful Life", "Revision Effective Period",                     # Y Z
    "Tax Book Differs Memo",                                                 # AA
    "Net Book Value (calc)", "Status (calc)",                                # AB AC
    "Significant Asset? (calc)", "Matches Policy? (calc)", "Notes",          # AD AE AF
    # v1.1 additions (appended; existing letters A..AF unchanged so all
    # downstream formulas referencing A..AF remain stable):
    "LHI Renewal Term (months)",   # AG (33) — read by Effective Useful Life when W="Y"
    "Voided? (Y/N)",               # AH (34) — soft-delete; gates Status/SUMIFS
]
# Columns A..AH = 34 columns. Sanity check
assert len(REGISTER_HEADERS) == 34, f"Register columns must be 34, got {len(REGISTER_HEADERS)}"

# input/calc map by column index (1-based)
REG_CALC_COLS = {21, 24, 28, 29, 30, 31}   # U=21, X=24, AB=28, AC=29, AD=30, AE=31
# unlocked input cols per spec: A-R (1-18), V (22), W (23), Y (25), Z (26), AA (27), AF (32),
# plus v1.1 additions AG (33) renewal-term, AH (34) voided
# S (19) is unlocked under separate password (Opening Accum Depr)
# T (20) is unlocked input
REG_INPUT_COLS = set(range(1, 19)) | {20, 22, 23, 25, 26, 27, 32, 33, 34}
REG_OPENING_COL = 19  # Opening Accum Depr — locked except with separate password


def build_register(wb: Workbook, variant: str):
    ws = wb.create_sheet("Asset Register")
    ws.sheet_properties.tabColor = "FF9BC2E6"
    max_assets = MAX_ASSETS[variant]

    # Column widths
    widths = {
        "A": 12, "B": 28, "C": 22, "D": 14, "E": 18, "F": 12, "G": 12,
        "H": 14, "I": 12, "J": 12, "K": 12, "L": 12, "M": 12, "N": 12,
        "O": 14, "P": 12, "Q": 14, "R": 14, "S": 14, "T": 12, "U": 12,
        "V": 14, "W": 14, "X": 14, "Y": 12, "Z": 14, "AA": 28,
        "AB": 14, "AC": 12, "AD": 14, "AE": 14, "AF": 30,
        "AG": 14, "AH": 10,  # v1.1: Renewal Term (months), Voided?
    }
    set_col_widths(ws, widths)

    # Header row
    for j, h in enumerate(REGISTER_HEADERS, start=1):
        c = ws.cell(row=1, column=j, value=h)
        style_header(c)
    ws.row_dimensions[1].height = 32

    # Examples + blank rows
    examples = EXAMPLES_PAID if variant == "paid" else EXAMPLES_FREE
    if variant == "free":
        # Drop UoP example if any (filter out method UoP)
        examples = [e for e in examples if e[14] != "UoP"]
        # Ensure 2 example rows
        examples = examples[:2]

    n_examples = len(examples)

    for i in range(max_assets):
        r = 2 + i
        is_example = i < n_examples
        fill = EXAMPLE_FILL if is_example else None
        ex = examples[i] if is_example else None

        # Set cell values
        def put(col_idx, value, number_format=None):
            cell = ws.cell(row=r, column=col_idx, value=value)
            if number_format:
                cell.number_format = number_format
            if fill:
                cell.fill = fill
            cell.border = BORDER_ALL
            return cell

        if is_example:
            # Map of EXAMPLES tuple → cols A..AF
            (aid, desc, cat, dept, vendor, inv, po, appr, appr_dt,
             acq, ins, cost, salv, life, method, ddb_rate, total_units, conv,
             open_accum, open_per, lhi_end, lhi_renew, rev_life, rev_per, tax_memo, notes) = ex

            put(1, aid)
            put(2, desc)
            put(3, cat)
            put(4, dept)
            put(5, vendor)
            put(6, inv)
            put(7, po)
            put(8, appr)
            put(9, appr_dt, "yyyy-mm-dd")
            put(10, acq, "yyyy-mm-dd")
            put(11, ins, "yyyy-mm-dd")
            put(12, cost, "#,##0.00")
            put(13, salv, "#,##0.00")
            put(14, life)
            put(15, method)
            put(16, ddb_rate, "0.00%")
            put(17, total_units)
            put(18, conv)
            put(19, open_accum, "#,##0.00")
            put(20, open_per, "yyyy-mm-dd")
            put(22, lhi_end, "yyyy-mm-dd")
            put(23, lhi_renew)
            put(25, rev_life)
            put(26, rev_per, "yyyy-mm-dd")
            put(27, tax_memo)
            put(32, notes)
            # v1.1 inputs: Renewal Term (months) for LHI; Voided?
            # FA-0003 is the LHI example — give it a sample renewal term of 60 months
            # (mirrors a "5-year renewal reasonably certain" scenario) only when LHI Renewal=Y.
            if cat == "Leasehold Improvements" and lhi_renew == "Y":
                put(33, 60)  # 5-year reasonably-certain renewal
            else:
                put(33, 0)
            put(34, "N")  # not voided
        else:
            # Asset ID auto-suggestion formula (user can override)
            put(1, f'=IF(LEN(B{r})>0,"FA-"&TEXT(ROW()-1,"0000"),"")')
            put(19, 0, "#,##0.00")
            put(20, "=Setup!$B$3", "yyyy-mm-dd")  # Opening Period default = FY start
            put(18, "=Setup!$B$8")  # Convention default
            put(33, 0)               # default renewal-term months
            put(34, "N")             # default not voided

        # Calc columns — always present
        # U = Depr Start Date = MAX(In-Service, Opening Period)
        put(21, f"=IF(LEN(A{r})=0,\"\",MAX(K{r},T{r}))", "yyyy-mm-dd")
        # X = Effective Useful Life =
        #     IF(Category="LHI", MIN(Life, months from In-Service to lease-end
        #         + (if Renewal=Y, add AG renewal-term months)),
        #     else Life)
        # ASC 842-20-35-12: LHI life capped at lease term INCLUDING reasonably-certain renewals.
        # Renewal handling added in v1.1: column AG holds renewal-term months; col W gates whether
        # the renewal is reasonably certain (Y/N).
        eff_life = (
            f'=IF(LEN(A{r})=0,"",'
            f'IF(C{r}="Leasehold Improvements",'
            f'MIN(N{r}, MAX(1,(YEAR(V{r})-YEAR(K{r}))*12+MONTH(V{r})-MONTH(K{r})+1'
            f'+IF(W{r}="Y",IF(ISNUMBER(AG{r}),AG{r},0),0))),'
            f'N{r}))'
        )
        put(24, eff_life)
        # AB = NBV = Cost - SUMIFS(Schedule.H where Asset==A) - SUMIFS(Disposal.NBV where Asset==A)
        # We'll use SUMIFS over the Schedule for total final depreciation; subtract Opening Accum included via running.
        # NBV = Cost - (Opening Accum + Schedule depr through reporting period) - (any disposal write-off proceeds adjustment NOT applied here — disposal handling reduces NBV by removing future depr; we treat NBV as Cost - Opening - depr-to-date if not Disposed, else 0)
        nbv = (
            f'=IF(LEN(A{r})=0,"",'
            f'IF(IFERROR(MATCH(A{r},Disposal_AssetID,0),0)>0,0,'
            f'L{r}-S{r}-SUMIFS(Sched_DeprFinal,Sched_AssetID_Col,A{r},Sched_Period_Col,"<="&Setup!$B$4)))'
        )
        put(28, nbv, "#,##0.00")
        # AC = Status
        # v1.1: now reads Voided? input (AH) first. Order:
        #   1) Voided?=Y  -> "Voided"
        #   2) In Disposal Log -> "Disposed"
        #   3) NBV <= Salvage -> "Fully Depreciated"
        #   4) Else "Active"
        status = (
            f'=IF(LEN(A{r})=0,"",'
            f'IF(AH{r}="Y","Voided",'
            f'IF(IFERROR(MATCH(A{r},Disposal_AssetID,0),0)>0,"Disposed",'
            f'IF(AB{r}<=M{r}+0.01,"Fully Depreciated","Active"))))'
        )
        put(29, status)
        # AD = Significant Asset?
        # v1.1: Voided assets never flag SIGNIFICANT (they are excluded from the active book).
        sig = (
            f'=IF(LEN(A{r})=0,"",'
            f'IF(AH{r}="Y","",'
            f'IF(AB{r}>Setup!$B$5,"SIGNIFICANT","")))'
        )
        put(30, sig)
        # AE = Matches Policy?
        # Compare N (Life) and O (Method) against Policy table by Category
        # Policy lives on Setup: cols A (cat), B (life), C (method), D (salvage)
        # Use INDEX/MATCH
        pol_match = (
            f'=IF(LEN(A{r})=0,"",'
            f'IF(AND('
            f'N{r}=INDEX(Policy_Life,MATCH(C{r},Policy_Category,0)),'
            f'O{r}=INDEX(Policy_Method,MATCH(C{r},Policy_Category,0))'
            f'),"Y","N"))'
        )
        put(31, pol_match)

        # Apply protection per column for this row
        for col_idx in range(1, 35):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = BORDER_ALL
            if col_idx in REG_CALC_COLS:
                cell.protection = LOCKED
            elif col_idx == REG_OPENING_COL:
                # Opening Accum Depr — separate password (we lock it; user must unprotect with sub-password)
                cell.protection = LOCKED
            elif col_idx in REG_INPUT_COLS:
                cell.protection = UNLOCKED
                if not fill:
                    cell.fill = INPUT_FILL
            else:
                cell.protection = LOCKED

        # Re-apply gray fill OVER input pale-yellow for example rows
        if fill:
            for col_idx in range(1, 35):
                ws.cell(row=r, column=col_idx).fill = fill

    # Data validations
    # Asset ID uniqueness
    end_row = 1 + max_assets
    dv_id = DataValidation(
        type="custom",
        formula1=f'=AND(LEN(A2)>0,COUNTIF($A$2:$A${end_row},A2)<=1)',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Duplicate Asset ID",
        error="Asset ID must be unique.",
    )
    dv_id.add(f"A2:A{end_row}")
    ws.add_data_validation(dv_id)

    dv_cat = DataValidation(type="list", formula1="=CategoryList", allow_blank=True)
    dv_cat.add(f"C2:C{end_row}")
    ws.add_data_validation(dv_cat)

    dv_method = DataValidation(type="list", formula1="=MethodList", allow_blank=True)
    dv_method.add(f"O2:O{end_row}")
    ws.add_data_validation(dv_method)

    dv_conv = DataValidation(type="list", formula1="=ConventionList", allow_blank=True)
    dv_conv.add(f"R2:R{end_row}")
    ws.add_data_validation(dv_conv)

    dv_lhi = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    dv_lhi.add(f"W2:W{end_row}")
    ws.add_data_validation(dv_lhi)

    dv_cost = DataValidation(
        type="custom",
        formula1=f"=L2>0",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Cost must be positive",
        error="Cost must be > 0.",
    )
    dv_cost.add(f"L2:L{end_row}")
    ws.add_data_validation(dv_cost)

    dv_salv = DataValidation(
        type="custom",
        formula1=f"=AND(M2>=0,M2<=L2)",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Salvage out of range",
        error="Salvage must satisfy 0 <= Salvage <= Cost.",
    )
    dv_salv.add(f"M2:M{end_row}")
    ws.add_data_validation(dv_salv)

    dv_life = DataValidation(
        type="custom",
        formula1=f"=AND(N2>=1,N2<=600,N2=INT(N2))",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Useful Life out of range",
        error="Useful life must be an integer between 1 and 600 months.",
    )
    dv_life.add(f"N2:N{end_row}")
    ws.add_data_validation(dv_life)

    dv_acq = DataValidation(
        type="custom",
        formula1=f"=AND(ISNUMBER(J2),J2<=Setup!$B$4)",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Acquisition Date",
        error="Acquisition Date must be <= Reporting Period End.",
    )
    dv_acq.add(f"J2:J{end_row}")
    ws.add_data_validation(dv_acq)

    dv_ins = DataValidation(
        type="custom",
        formula1=f"=AND(ISNUMBER(K2),K2>=J2)",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="In-Service before Acquisition",
        error="In-Service Date must be on or after Acquisition Date.",
    )
    dv_ins.add(f"K2:K{end_row}")
    ws.add_data_validation(dv_ins)

    # v1.1: Renewal Term months (AG) — integer 0..600
    dv_renew = DataValidation(
        type="custom",
        formula1=f"=AND(ISNUMBER(AG2),AG2>=0,AG2<=600,AG2=INT(AG2))",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Renewal Term out of range",
        error="Renewal Term must be an integer between 0 and 600 months.",
    )
    dv_renew.add(f"AG2:AG{end_row}")
    ws.add_data_validation(dv_renew)

    # v1.1: Voided? (AH) — Y/N dropdown
    dv_void = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    dv_void.add(f"AH2:AH{end_row}")
    ws.add_data_validation(dv_void)

    # Freeze header
    ws.freeze_panes = "C2"

    ws.protection.set_password(SHEET_PW)
    ws.protection.sheet = True

    ws._anchors = {"end_row": end_row}


# ---------------------------------------------------------------------------
# Tab 3: Schedule
# ---------------------------------------------------------------------------

SCHED_HEADERS_BASE = [
    "Asset ID",                          # A
    "Period (Month End)",                # B
    "Period Index",                      # C
    "Beginning NBV",                     # D
    "Method (locked)",                   # E
    "Depr This Period (raw)",            # F
    "Disposal Adjustment",               # G
    "Depr This Period (final)",          # H
    "Accumulated Depreciation",          # I
    "Ending NBV",                        # J
    "Asset Match Idx",                   # K (Sched_AssetIdx)
    "Category Period Key",               # L
    "JE Reference",                      # M
    "In Closed Period?",                 # N
    "Switch Flag (DDB)",                 # O
]
# v1.1: helper columns for Cat×Dept aggregation (P0-4) and Voided filter (P0-3).
# Free: no Units Produced col, so helpers land at P, Q, R.
# Paid: Units Produced at P, helpers at Q, R, S.
# In both variants the helpers always occupy positions:
#   - Department (calc, from Register col D)
#   - Cat-Dept-Period Key (calc, "Cat|Dept|YYYY-MM")
#   - Voided? (calc, from Register col AH)
SCHED_HEADERS_PAID = SCHED_HEADERS_BASE + [
    "Units Produced This Month",         # P
    "Department (calc)",                 # Q
    "Cat-Dept-Period Key (calc)",        # R
    "Voided? (calc)",                    # S
]
SCHED_HEADERS_FREE = SCHED_HEADERS_BASE + [
    "Department (calc)",                 # P
    "Cat-Dept-Period Key (calc)",        # Q
    "Voided? (calc)",                    # R
]


def build_schedule(wb: Workbook, variant: str):
    ws = wb.create_sheet("Schedule")
    ws.sheet_properties.tabColor = "FFA9D08E"

    max_assets = MAX_ASSETS[variant]
    depth = SCHED_DEPTH[variant]
    headers = SCHED_HEADERS_PAID if variant == "paid" else SCHED_HEADERS_FREE
    n_cols = len(headers)

    # v1.1: column letters for new helper cols differ by variant
    # paid: Q=Dept, R=CatDeptKey, S=Voided ; free: P=Dept, Q=CatDeptKey, R=Voided
    if variant == "paid":
        DEPT_COL_IDX, CDK_COL_IDX, VOID_COL_IDX = 17, 18, 19
        DEPT_COL_LETTER, CDK_COL_LETTER, VOID_COL_LETTER = "Q", "R", "S"
    else:
        DEPT_COL_IDX, CDK_COL_IDX, VOID_COL_IDX = 16, 17, 18
        DEPT_COL_LETTER, CDK_COL_LETTER, VOID_COL_LETTER = "P", "Q", "R"

    # Column widths
    widths = {
        "A": 12, "B": 14, "C": 10, "D": 14, "E": 14, "F": 14, "G": 14,
        "H": 14, "I": 14, "J": 14, "K": 10, "L": 24, "M": 18, "N": 14, "O": 10,
        "P": 14, "Q": 18, "R": 28, "S": 10,
    }
    set_col_widths(ws, widths)

    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        style_header(c)
    ws.row_dimensions[1].height = 32

    # Generate rows: max_assets * depth
    # For row r in 2..1+max_assets*depth:
    #   asset_idx (0..max_assets-1) = (r-2) // depth
    #   period_idx (1..depth) = (r-2) % depth + 1
    total_rows = max_assets * depth

    # Pre-compute first-row-of-asset boolean by formula via A<>A previous; but our pattern
    # uses Beginning NBV with reference to G prior row. We have ample formulas; let's do it efficient.

    for row_offset in range(total_rows):
        r = 2 + row_offset
        asset_idx = row_offset // depth
        period_idx = (row_offset % depth) + 1
        reg_row = 2 + asset_idx  # 1-based register row

        # A: Asset ID = ='Asset Register'!A{reg_row}  (only if present)
        ws.cell(row=r, column=1, value=f"='Asset Register'!A{reg_row}")
        # K: Asset Match Idx — since 1:1 mapping, K = asset_idx + 1 (the row index within Register). But
        # spec wants =MATCH(A2, Register.A, 0). We hardcode the constant for performance + because A column
        # references register directly. We'll use the constant to keep formula count down.
        ws.cell(row=r, column=11, value=asset_idx + 1)

        # B: Period (Month End) = EOMONTH(Depr Start Date, Period Index - 1)
        # Depr Start Date = Register col U
        ws.cell(row=r, column=2, value=f'=IF(LEN(A{r})=0,"",EOMONTH(\'Asset Register\'!U{reg_row},{period_idx - 1}))').number_format = "yyyy-mm-dd"
        # C: Period Index
        ws.cell(row=r, column=3, value=period_idx)

        # D: Beginning NBV — Option B
        # If period_idx == 1: Cost - Opening Accum Depr
        # Else: prior row Ending NBV (J)
        if period_idx == 1:
            ws.cell(row=r, column=4, value=f'=IF(LEN(A{r})=0,"",\'Asset Register\'!L{reg_row}-\'Asset Register\'!S{reg_row})').number_format = "#,##0.00"
        else:
            ws.cell(row=r, column=4, value=f'=IF(LEN(A{r})=0,"",J{r - 1})').number_format = "#,##0.00"

        # E: Method (locked) = INDEX(Register.O, K)
        ws.cell(row=r, column=5, value=f"=\'Asset Register\'!O{reg_row}")

        # O: Switch Flag (DDB)
        # Once SL on remaining ≥ DDB in any period, "SL" forever after; else "DDB"
        # SL_rem = (D - Salvage) / max(1, EffLife - PeriodIdx + 1)
        # DDB_rate_monthly = (DDBOverride or 2/(EffLife/12))/12 ; we keep override on Register P; default 2/years
        # Use: DDB_amt = D * effective_rate / 12 where rate = IF(P_blank, 2/(X/12), P)
        # Compare SL_rem to DDB_amt
        # Use Register col P for override, X for effective life
        eff_life_ref = f"'Asset Register'!X{reg_row}"
        sl_amt_expr = f"IF({eff_life_ref}-C{r}+1>0,(D{r}-'Asset Register'!M{reg_row})/({eff_life_ref}-C{r}+1),0)"
        ddb_rate_expr = f"IF(LEN('Asset Register'!P{reg_row})>0,'Asset Register'!P{reg_row},IF({eff_life_ref}>0,2/({eff_life_ref}/12),0))"
        ddb_amt_expr = f"D{r}*({ddb_rate_expr})/12"
        if period_idx == 1:
            sf_formula = (
                f'=IF(LEN(A{r})=0,"",IF(E{r}<>"DDB","N/A",'
                f'IF({sl_amt_expr}>={ddb_amt_expr},"SL","DDB")))'
            )
        else:
            sf_formula = (
                f'=IF(LEN(A{r})=0,"",IF(E{r}<>"DDB","N/A",'
                f'IF(OR(O{r - 1}="SL",{sl_amt_expr}>={ddb_amt_expr}),"SL","DDB")))'
            )
        ws.cell(row=r, column=15, value=sf_formula)

        # F: Depr This Period (raw)
        # Branch by Method (E):
        #   SL  -> (Cost - Salvage)/EffLife
        #   DDB -> IF switch=SL -> SL_amt, else MAX(DDB_amt, SL_amt)
        #   SYD -> closed form below
        #   UoP -> (Cost-Salvage)/TotalUnits * UnitsProducedThisMonth  (only paid; col P present)
        cost_ref = f"'Asset Register'!L{reg_row}"
        salv_ref = f"'Asset Register'!M{reg_row}"
        life_ref = eff_life_ref  # use effective life including LHI cap
        total_units_ref = f"'Asset Register'!Q{reg_row}"

        # Years and SYD anchor — SYD relative to In-Service Date (we anchor period index from Depr Start)
        # LifeYrs = X/12
        # SumDigits = LifeYrs*(LifeYrs+1)/2
        # YearK = INT((C-1)/12)+1
        # MonthInYr = MOD(C-1,12)+1
        # Frac1 = (LifeYrs - YearK + 1) / SumDigits
        # Frac2 = (LifeYrs - YearK) / SumDigits  (next year)
        # Raw = (Cost-Salvage) * (Frac1*(13-MonthInYr) + Frac2*(MonthInYr-1)) / 12
        sl_raw = f"({cost_ref}-{salv_ref})/MAX(1,{life_ref})"

        ddb_raw = f"IF(O{r}=\"SL\",{sl_amt_expr},MAX({ddb_amt_expr},{sl_amt_expr}))"

        syd_block = (
            f"(({cost_ref}-{salv_ref})*("
            f"(({life_ref}/12)-INT((C{r}-1)/12)-1+1)/(({life_ref}/12)*(({life_ref}/12)+1)/2)"
            f"*(13-MOD(C{r}-1,12)-1)"
            f"+(({life_ref}/12)-INT((C{r}-1)/12)-1)/(({life_ref}/12)*(({life_ref}/12)+1)/2)"
            f"*(MOD(C{r}-1,12)+1-1)"
            f")/12)"
        )

        # P1-7 fix: Mid-Month convention proration on first period of capitalization.
        # When C=1 AND Convention="Mid-Month", multiply the raw time-based depr by
        # (DAY(in-service) - 0.5) / DAY(EOM(in-service)). UoP excluded — units-based
        # methods prorate via the units-produced input itself.
        in_svc_ref = f"'Asset Register'!K{reg_row}"
        conv_ref = f"'Asset Register'!R{reg_row}"
        first_pd_factor = (
            f"IF(AND(C{r}=1,{conv_ref}=\"Mid-Month\"),"
            f"(DAY({in_svc_ref})-0.5)/DAY(EOMONTH({in_svc_ref},0)),1)"
        )
        if variant == "paid":
            uop_per_unit = f"IF({total_units_ref}>0,({cost_ref}-{salv_ref})/{total_units_ref},0)"
            uop_raw = f"({uop_per_unit})*IF(LEN(P{r})>0,P{r},0)"
            # Time-based methods (SL/DDB/SYD) get the first-period factor; UoP does not.
            f_formula = (
                f'=IF(LEN(A{r})=0,"",'
                f'IF(E{r}="Straight-Line",({sl_raw})*({first_pd_factor}),'
                f'IF(E{r}="DDB",({ddb_raw})*({first_pd_factor}),'
                f'IF(E{r}="SYD",({syd_block})*({first_pd_factor}),'
                f'IF(E{r}="UoP",{uop_raw},0)))))'
            )
        else:
            f_formula = (
                f'=IF(LEN(A{r})=0,"",'
                f'IF(E{r}="Straight-Line",({sl_raw})*({first_pd_factor}),'
                f'IF(E{r}="DDB",({ddb_raw})*({first_pd_factor}),'
                f'IF(E{r}="SYD",({syd_block})*({first_pd_factor}),0))))'
            )
        ws.cell(row=r, column=6, value=f_formula).number_format = "#,##0.00"

        # G: Disposal Adjustment
        # If asset not in Disposal Log, 0.
        # If asset disposed and current period == EOMONTH(disp date), prorate per convention:
        #   Mid-Month: F * ((DAY(disp) - 0.5) / DAY(B)) (effectively keep depr through disp day) — adjustment shrinks; we set G to amount that REDUCES F to that prorated amount.
        #   Full Month: 0 (full month already)
        # Adjustment definition: H = F + G. We want H = full F if Full Month; H = F * frac if Mid-Month.
        # So G = F * (frac - 1) when Mid-Month, else 0.
        # If after disposal month, zero out depr -> G = -F.
        disp_date = f"IFERROR(INDEX(Disposal_Date,MATCH(A{r},Disposal_AssetID,0)),0)"
        conv_ref = f"'Asset Register'!R{reg_row}"
        g_formula = (
            f'=IF(LEN(A{r})=0,"",'
            f'IF({disp_date}=0,0,'
            f'IF(B{r}<EOMONTH({disp_date},0),0,'
            f'IF(B{r}>EOMONTH({disp_date},0),-F{r},'
            f'IF({conv_ref}="Mid-Month",F{r}*((DAY({disp_date})-0.5)/DAY(B{r})-1),0)'
            f'))))'
        )
        ws.cell(row=r, column=7, value=g_formula).number_format = "#,##0.00"

        # H: Final Depr — salvage floor wrapper
        # H = MIN(MAX(0, D - Salvage), F + G)
        h_formula = (
            f'=IF(LEN(A{r})=0,"",MIN(MAX(0,D{r}-\'Asset Register\'!M{reg_row}),F{r}+G{r}))'
        )
        ws.cell(row=r, column=8, value=h_formula).number_format = "#,##0.00"

        # I: Accumulated Depreciation = running sum since asset start
        if period_idx == 1:
            ws.cell(row=r, column=9, value=f"=IF(LEN(A{r})=0,\"\",\'Asset Register\'!S{reg_row}+H{r})").number_format = "#,##0.00"
        else:
            ws.cell(row=r, column=9, value=f'=IF(LEN(A{r})=0,"",I{r - 1}+H{r})').number_format = "#,##0.00"

        # J: Ending NBV = D - H
        ws.cell(row=r, column=10, value=f'=IF(LEN(A{r})=0,"",D{r}-H{r})').number_format = "#,##0.00"

        # L: Category Period Key = Category & "|" & TEXT(Period, "YYYY-MM")
        ws.cell(row=r, column=12, value=f'=IF(LEN(A{r})=0,"",\'Asset Register\'!C{reg_row}&"|"&TEXT(B{r},"YYYY-MM"))')

        # M: JE Reference = A & "-" & TEXT(B,"YYYY-MM")
        ws.cell(row=r, column=13, value=f'=IF(LEN(A{r})=0,"",A{r}&"-"&TEXT(B{r},"YYYY-MM"))')

        # N: In Closed Period?
        # COUNTIFS over Setup!ClosedPeriods (col A) for entries <= B; if >0, "LOCKED"
        anchors = wb["Setup"]._anchors
        cp_start = anchors["ClosedPeriods_Start_Row"]
        cp_end = anchors["ClosedPeriods_End_Row"]
        ws.cell(row=r, column=14, value=(
            f'=IF(LEN(A{r})=0,"",IF(COUNTIFS(Setup!$A${cp_start}:$A${cp_end},B{r})>0,"LOCKED",""))'
        ))

        # P: Units Produced This Month (paid only) — UNLOCKED input cell when UoP
        if variant == "paid":
            cell = ws.cell(row=r, column=16, value=None)
            cell.number_format = "#,##0"
            # Unlock if Method = UoP — we can't conditionally protect via formula, so unlock all P cells.
            cell.protection = UNLOCKED
            cell.fill = INPUT_FILL

        # v1.1 helper columns (always populated; positions differ by variant)
        # Department (calc) — pulls Register col D ("" if blank so SUMIFS criteria match "" reliably)
        ws.cell(row=r, column=DEPT_COL_IDX, value=(
            f'=IF(LEN(A{r})=0,"",IF(LEN(\'Asset Register\'!D{reg_row})>0,\'Asset Register\'!D{reg_row},""))'
        ))
        # Cat-Dept-Period Key (calc) = "Cat|Dept|YYYY-MM"
        ws.cell(row=r, column=CDK_COL_IDX, value=(
            f'=IF(LEN(A{r})=0,"",\'Asset Register\'!C{reg_row}&"|"&'
            f'{DEPT_COL_LETTER}{r}&"|"&TEXT(B{r},"YYYY-MM"))'
        ))
        # Voided? (calc) — mirrors Register col AH ("Y"/"N")
        ws.cell(row=r, column=VOID_COL_IDX, value=(
            f'=IF(LEN(A{r})=0,"",IFERROR(\'Asset Register\'!AH{reg_row},"N"))'
        ))

        # Protect all other cells in this row by default
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = BORDER_ALL
            if variant == "paid" and col_idx == 16:
                continue  # already set
            cell.protection = LOCKED

    # Hide helper columns K, O, helper-block (Dept/CDK/Voided)
    ws.column_dimensions["K"].hidden = True
    ws.column_dimensions["O"].hidden = True
    ws.column_dimensions[DEPT_COL_LETTER].hidden = True
    ws.column_dimensions[CDK_COL_LETTER].hidden = True
    ws.column_dimensions[VOID_COL_LETTER].hidden = True

    # Conditional formatting: closed period rows -> gray fill, italic
    closed_rule = FormulaRule(
        formula=[f'$N2="LOCKED"'],
        fill=PatternFill(fill_type="solid", start_color="FFE0E0E0", end_color="FFE0E0E0"),
        font=Font(italic=True, color="FF888888"),
    )
    ws.conditional_formatting.add(f"A2:{get_column_letter(n_cols)}{1 + total_rows}", closed_rule)

    ws.freeze_panes = "D2"
    ws.protection.set_password(SHEET_PW)
    ws.protection.sheet = True

    ws._anchors = {
        "end_row": 1 + total_rows,
        "n_cols": n_cols,
        "total_rows": total_rows,
        "DEPT_COL_LETTER": DEPT_COL_LETTER,
        "CDK_COL_LETTER": CDK_COL_LETTER,
        "VOID_COL_LETTER": VOID_COL_LETTER,
    }


# ---------------------------------------------------------------------------
# Tab 4: JE Generator
# ---------------------------------------------------------------------------

def build_je_generator(wb: Workbook, variant: str):
    ws = wb.create_sheet("JE Generator")
    ws.sheet_properties.tabColor = "FFFFE699"

    set_col_widths(ws, {
        "A": 12, "B": 28, "C": 36, "D": 14, "E": 14, "F": 22, "G": 16, "H": 28,
        "I": 16, "J": 16, "K": 16,
    })

    # Top: show selected period + system
    ws["A1"] = "Reporting Period:"
    ws["B1"] = "=Setup!$B$4"
    ws["B1"].number_format = "yyyy-mm-dd"
    ws["A2"] = "GL System Preset:"
    ws["B2"] = "=Setup!$B$11" if variant == "paid" else "Generic"

    # Header row at row 4
    base_headers = ["Date", "Account", "Description", "Debit", "Credit", "Reference", "Department", "Memo"]

    if variant == "paid":
        # Per-system extra columns: build a row of dynamic headers via IF on GLSystem.
        # Columns I, J, K hold the system-specific headers/values.
        extra_headers_formula = [
            ('I4', '=IF($B$2="NetSuite","Subsidiary",IF($B$2="QuickBooks","Class",IF($B$2="Sage Intacct","Department",IF($B$2="Xero","Tracking Category 1",""))))'),
            ('J4', '=IF($B$2="NetSuite","Class",IF($B$2="Sage Intacct","Location",IF($B$2="Xero","Tracking Category 2","")))'),
            ('K4', '=IF($B$2="NetSuite","Location","")'),
        ]
    else:
        extra_headers_formula = []

    for j, h in enumerate(base_headers, start=1):
        c = ws.cell(row=4, column=j, value=h)
        style_header(c)
    if variant == "paid":
        for cell_ref, formula in extra_headers_formula:
            c = ws[cell_ref]
            c.value = formula
            style_header(c)

    ws.row_dimensions[4].height = 32

    # Aggregated depreciation lines by Category (one DR per category in the reporting period, one CR Accum Depr aggregate)
    # We'll emit:
    #  - 6 category DR lines (Depreciation Expense by category, by department blank/aggregate for V1)
    #  - 1 CR Accum Depr line
    #  - Capitalization lines (one per category) for assets in service this period
    #  - Disposal-related lines (variable; emit fixed slots and let users zero blanks)

    # For simplicity and audit clarity, emit:
    # Row 5+ for monthly depreciation (one row per category)
    # Then Cap + Disposal rows below.

    row = 5
    # Depreciation expense DR by category
    for cat in CATEGORIES:
        # Date = ReportingPeriodEnd
        ws.cell(row=row, column=1, value="=Setup!$B$4").number_format = "yyyy-mm-dd"
        # Account: GL code label for Depr Expense + descriptor
        ws.cell(row=row, column=2, value=f'=Setup!$B$21&" - {cat}"')
        ws.cell(row=row, column=3, value=f"Monthly depreciation — {cat}")
        # Debit = SUMIFS(Sched_DeprFinal, Sched_Period, Reporting Period, Sched_Cat, cat) where Sched_Cat is calc as Category period key startswith cat
        # We'll use sum on Sched_DeprFinal with Sched_CatKey starts with cat|YYYY-MM
        cat_key = f'"{cat}|"&TEXT(Setup!$B$4,"YYYY-MM")'
        ws.cell(row=row, column=4, value=(
            f'=SUMIFS(Sched_DeprFinal,Sched_CatKey,{cat_key})'
        )).number_format = "#,##0.00"
        ws.cell(row=row, column=5, value=0).number_format = "#,##0.00"
        ws.cell(row=row, column=6, value=f'="DEPR-"&TEXT(Setup!$B$4,"YYYY-MM")&"-{cat[:3].upper()}"')
        ws.cell(row=row, column=7, value="")  # Department aggregated — V1 leaves blank
        ws.cell(row=row, column=8, value=f"Aggregated from Schedule — {cat}")
        row += 1

    # CR Accumulated Depreciation total
    ws.cell(row=row, column=1, value="=Setup!$B$4").number_format = "yyyy-mm-dd"
    ws.cell(row=row, column=2, value="=Setup!$B$20")
    ws.cell(row=row, column=3, value="Monthly depreciation — Accumulated Depreciation")
    ws.cell(row=row, column=4, value=0).number_format = "#,##0.00"
    # Credit = SUMIFS over all Sched_DeprFinal at this period
    ws.cell(row=row, column=5, value=(
        '=SUMIFS(Sched_DeprFinal,Sched_Period_Col,Setup!$B$4)'
    )).number_format = "#,##0.00"
    ws.cell(row=row, column=6, value='="DEPR-"&TEXT(Setup!$B$4,"YYYY-MM")&"-AD"')
    ws.cell(row=row, column=8, value="Aggregated CR to Accum Depr")
    row += 2

    # Capitalization rows — emit one slot per category for cap of assets in service this period
    ws.cell(row=row, column=2, value="— Capitalization (assets placed in service this period) —").font = SECTION_FONT
    row += 1
    reg_end = wb["Asset Register"]._anchors["end_row"]
    for cat in CATEGORIES:
        ws.cell(row=row, column=1, value="=Setup!$B$4").number_format = "yyyy-mm-dd"
        # Account = GL category code
        gl_code = {
            "Computer Equipment": "$B$14",
            "Furniture & Fixtures": "$B$15",
            "Leasehold Improvements": "$B$16",
            "Machinery": "$B$17",
            "Vehicles": "$B$18",
            "Other": "$B$19",
        }[cat]
        ws.cell(row=row, column=2, value=f'=Setup!{gl_code}&" - {cat}"')
        ws.cell(row=row, column=3, value=f"Capitalization — {cat}")
        # Debit = SUMIFS(Register.L, Register.K (InService) >= start of period, <= end of period, Register.C = cat)
        # We'll use SUMIFS with Register_Cost, Register_InServiceDate >= EOMONTH(ReportingPeriod,-1)+1 and <= ReportingPeriod, and Register_Category = cat
        debit_formula = (
            f'=SUMIFS(Register_Cost,Register_Category,"{cat}",'
            f'Register_InServiceDate,">="&(EOMONTH(Setup!$B$4,-1)+1),'
            f'Register_InServiceDate,"<="&Setup!$B$4)'
        )
        ws.cell(row=row, column=4, value=debit_formula).number_format = "#,##0.00"
        ws.cell(row=row, column=5, value=0).number_format = "#,##0.00"
        ws.cell(row=row, column=6, value=f'="CAP-"&TEXT(Setup!$B$4,"YYYY-MM")&"-{cat[:3].upper()}"')
        row += 1

    # CR Cash/AP for capitalization
    ws.cell(row=row, column=1, value="=Setup!$B$4").number_format = "yyyy-mm-dd"
    ws.cell(row=row, column=2, value="Cash / AP")
    ws.cell(row=row, column=3, value="Capitalization — credit (cash or AP)")
    ws.cell(row=row, column=4, value=0).number_format = "#,##0.00"
    # Credit = SUMIFS Register_Cost with InServiceDate in period
    ws.cell(row=row, column=5, value=(
        '=SUMIFS(Register_Cost,Register_InServiceDate,">="&(EOMONTH(Setup!$B$4,-1)+1),Register_InServiceDate,"<="&Setup!$B$4)'
    )).number_format = "#,##0.00"
    ws.cell(row=row, column=6, value='="CAP-"&TEXT(Setup!$B$4,"YYYY-MM")&"-CASH"')
    row += 2

    # Disposal section — pull from Disposal Log; emit up to 25 rows of disposal-pair lines
    ws.cell(row=row, column=2, value="— Disposals (Sale / Retirement / Trade-in / Write-off / Theft-loss) —").font = SECTION_FONT
    row += 1
    # Header row clarifying disposal lines
    # For each Disposal Log row in reporting period, emit:
    #   DR Cash (proceeds), DR Accum Depr, DR Loss OR CR Gain, CR Asset (at cost)
    # We'll emit one composite line per disposal slot, plus separately the gain/loss accounts.

    disp_max = 25 if variant == "paid" else 5
    for i in range(disp_max):
        disp_row = 2 + i  # Disposal Log row index (1-based with header on row 1)
        # P0-2 fix: gate every disposal cell on (a) Disposal Log row populated AND
        # (b) disposal-date month equals Reporting Period (Setup!$B$4). Disposals from
        # other periods render as "" so they don't leak into the JE export.
        gate = (f'OR(LEN(\'Disposal Log\'!A{disp_row})=0,'
                f'EOMONTH(\'Disposal Log\'!B{disp_row},0)<>Setup!$B$4)')
        # DR Cash
        ws.cell(row=row, column=1, value=f'=IF({gate},"",\'Disposal Log\'!B{disp_row})').number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=2, value=f'=IF({gate},"","Cash / AR")')
        ws.cell(row=row, column=3, value=f'=IF({gate},"","Disposal — DR Cash (proceeds) for "&\'Disposal Log\'!A{disp_row})')
        ws.cell(row=row, column=4, value=f'=IF({gate},"",\'Disposal Log\'!E{disp_row})').number_format = "#,##0.00"
        ws.cell(row=row, column=5, value="")
        ws.cell(row=row, column=6, value=f'=IF({gate},"","DISP-"&TEXT(\'Disposal Log\'!B{disp_row},"YYYY-MM")&"-"&\'Disposal Log\'!A{disp_row})')
        row += 1
        # DR Accum Depr
        ws.cell(row=row, column=1, value=f'=IF({gate},"",\'Disposal Log\'!B{disp_row})').number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=2, value=f'=IF({gate},"",Setup!$B$20)')
        ws.cell(row=row, column=3, value=f'=IF({gate},"","Disposal — DR Accum Depr for "&\'Disposal Log\'!A{disp_row})')
        ws.cell(row=row, column=4, value=f'=IF({gate},"",\'Disposal Log\'!L{disp_row})').number_format = "#,##0.00"
        ws.cell(row=row, column=5, value="")
        ws.cell(row=row, column=6, value=f'=IF({gate},"","DISP-"&TEXT(\'Disposal Log\'!B{disp_row},"YYYY-MM")&"-"&\'Disposal Log\'!A{disp_row})')
        row += 1
        # CR Asset (cost)
        ws.cell(row=row, column=1, value=f'=IF({gate},"",\'Disposal Log\'!B{disp_row})').number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=2, value=f'=IF({gate},"","Asset (at cost)")')
        ws.cell(row=row, column=3, value=f'=IF({gate},"","Disposal — CR Asset at cost for "&\'Disposal Log\'!A{disp_row})')
        ws.cell(row=row, column=4, value="")
        ws.cell(row=row, column=5, value=f'=IF({gate},"",\'Disposal Log\'!K{disp_row})').number_format = "#,##0.00"
        ws.cell(row=row, column=6, value=f'=IF({gate},"","DISP-"&TEXT(\'Disposal Log\'!B{disp_row},"YYYY-MM")&"-"&\'Disposal Log\'!A{disp_row})')
        row += 1
        # Gain/Loss line
        # If Gain (N > 0): CR Gain
        # If Loss (N < 0): DR Loss
        gl_acct = f'=IF({gate},"",IF(\'Disposal Log\'!N{disp_row}>0,Setup!$B$22,IF(\'Disposal Log\'!N{disp_row}<0,Setup!$B$23,"")))'
        gl_desc = f'=IF({gate},"",IF(\'Disposal Log\'!N{disp_row}>0,"Disposal — CR Gain for "&\'Disposal Log\'!A{disp_row},IF(\'Disposal Log\'!N{disp_row}<0,"Disposal — DR Loss for "&\'Disposal Log\'!A{disp_row},"")))'
        ws.cell(row=row, column=1, value=f'=IF({gate},"",\'Disposal Log\'!B{disp_row})').number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=2, value=gl_acct)
        ws.cell(row=row, column=3, value=gl_desc)
        # DR Loss (positive when N<0)
        ws.cell(row=row, column=4, value=f'=IF({gate},"",IF(\'Disposal Log\'!N{disp_row}<0,-\'Disposal Log\'!N{disp_row},""))').number_format = "#,##0.00"
        ws.cell(row=row, column=5, value=f'=IF({gate},"",IF(\'Disposal Log\'!N{disp_row}>0,\'Disposal Log\'!N{disp_row},""))').number_format = "#,##0.00"
        ws.cell(row=row, column=6, value=f'=IF({gate},"","DISP-"&TEXT(\'Disposal Log\'!B{disp_row},"YYYY-MM")&"-"&\'Disposal Log\'!A{disp_row})')
        row += 1

    # Borders for the printed area
    end_row = row - 1
    for r in range(4, end_row + 1):
        for c_idx in range(1, 12):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = BORDER_ALL
            cell.protection = LOCKED

    ws.freeze_panes = "A5"
    ws.protection.set_password(SHEET_PW)
    ws.protection.sheet = True
    ws._anchors = {"end_row": end_row}


# ---------------------------------------------------------------------------
# Tab 5: Rollforward
# ---------------------------------------------------------------------------

def build_rollforward(wb: Workbook, variant: str):
    ws = wb.create_sheet("Rollforward")
    ws.sheet_properties.tabColor = "FFFFD966"

    set_col_widths(ws, {"A": 26, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16})

    ws["A1"] = "Fixed Asset Rollforward — for Reporting Period ending"
    ws["A1"].font = TITLE_FONT
    ws["G1"] = "=Setup!$B$4"
    ws["G1"].number_format = "yyyy-mm-dd"

    # Header row
    cats = CATEGORIES + ["TOTAL"]
    for j, c in enumerate(cats, start=2):
        cell = ws.cell(row=3, column=j, value=c)
        style_header(cell)
    ws.cell(row=3, column=1, value="")
    style_header(ws.cell(row=3, column=1))

    # Period date refs
    rp_end_ref = "Setup!$B$4"
    fy_start_ref = "Setup!$B$3"
    # Opening period = FY start ; reporting period end = rp_end
    # For each category, compute:
    # Cost Opening: SUMIFS(Register_Cost, Register_Category=cat, Register_InServiceDate < FY_Start, Status != Disposed before FY_Start)
    #   Simplification: assets with InService < FY_Start AND not Disposed before FY_Start. We'll approximate: Disposed assets disposed before FY_Start excluded.
    # Cost Additions: Register_Cost where InService >= FY_Start AND InService <= RP_End AND Category = cat
    # Cost Disposals: -SUMIFS(Register_Cost, Register_Category=cat, MATCH in Disposal Log with disp date <= RP_End) using a different approach since SUMIFS can't directly match across tables — use SUMIFS on Disposal data we maintain on the Disposal Log:
    #   We'll add helper columns in Disposal Log that copy Category from Register so we can SUMIFS by Category.

    rows = [
        ("Cost", "Opening (before FY start)", "opening_cost"),
        ("", "Additions (in period)", "add_cost"),
        ("", "Disposals (in period)", "disp_cost"),
        ("", "Closing", "close_cost"),
        ("Accum Depr", "Opening", "opening_ad"),
        ("", "Depr Expense (period)", "depr_exp"),
        ("", "Disposals (in period)", "disp_ad"),
        ("", "Closing", "close_ad"),
        ("NBV", "Closing NBV", "nbv"),
    ]

    base_r = 4
    for ri, (group, label, key) in enumerate(rows):
        r = base_r + ri
        ws.cell(row=r, column=1, value=f"{group} — {label}" if group else label)
        ws.cell(row=r, column=1).font = Font(bold=bool(group))
        for j, cat in enumerate(cats, start=2):
            cell = ws.cell(row=r, column=j)
            if cat == "TOTAL":
                # Sum across category columns
                start_col = get_column_letter(2)
                end_col = get_column_letter(len(CATEGORIES) + 1)
                cell.value = f"=SUM({start_col}{r}:{end_col}{r})"
            else:
                cell.value = _rollforward_cell_formula(key, cat, rp_end_ref, fy_start_ref)
            cell.number_format = "#,##0.00"
            cell.border = BORDER_ALL
            cell.protection = LOCKED

        # Style the group row label
        ws.cell(row=r, column=1).border = BORDER_ALL

    ws.protection.set_password(SHEET_PW)
    ws.protection.sheet = True
    ws._anchors = {"base_r": base_r, "n_rows": len(rows)}


def _rollforward_cell_formula(key: str, cat: str, rp_end: str, fy_start: str) -> str:
    # Common SUMIFS templates
    if key == "opening_cost":
        # Cost of assets in service before FY start (and not disposed before FY start)
        return (
            f'=SUMIFS(Register_Cost,Register_Category,"{cat}",'
            f'Register_InServiceDate,"<"&({fy_start}))'
            f'-SUMIFS(Disposal_Cost,Disposal_Cat,"{cat}",Disposal_Date,"<"&({fy_start}))'
        )
    if key == "add_cost":
        return (
            f'=SUMIFS(Register_Cost,Register_Category,"{cat}",'
            f'Register_InServiceDate,">="&({fy_start}),'
            f'Register_InServiceDate,"<="&({rp_end}))'
        )
    if key == "disp_cost":
        return (
            f'=-SUMIFS(Disposal_Cost,Disposal_Cat,"{cat}",'
            f'Disposal_Date,">="&({fy_start}),Disposal_Date,"<="&({rp_end}))'
        )
    if key == "close_cost":
        # Opening + Additions + Disposals (disp_cost already signed)
        # Reference the cells: row offsets relative to base
        # We'll express closing using same SUMIFS for clarity:
        return (
            f'=SUMIFS(Register_Cost,Register_Category,"{cat}",'
            f'Register_InServiceDate,"<="&({rp_end}))'
            f'-SUMIFS(Disposal_Cost,Disposal_Cat,"{cat}",Disposal_Date,"<="&({rp_end}))'
        )
    if key == "opening_ad":
        # CatKey is text "{cat}|YYYY-MM"; SUMIFS criteria are lexical. We bound both sides
        # by the cat prefix so the sum stays inside this category — otherwise later-alphabet
        # cats over-match all earlier-alphabet rows. The ">="&cat&"|" floor is the empty-date
        # version, lexically lower than any cat|YYYY-MM key.
        return (
            f'=SUMIFS(Sched_DeprFinal,Sched_CatKey,">="&"{cat}|",'
            f'Sched_CatKey,"<="&"{cat}|"&TEXT(EOMONTH({fy_start},-1),"YYYY-MM"))'
            f'+SUMIFS(Register_Opening_AD,Register_Category,"{cat}")'
            f'-SUMIFS(Disposal_AD,Disposal_Cat,"{cat}",Disposal_Date,"<"&({fy_start}))'
        )
    if key == "depr_exp":
        # Depr expense in period: sum Sched_DeprFinal where Sched_CatKey >= cat|FY_start and <= cat|RP_end
        return (
            f'=SUMIFS(Sched_DeprFinal,Sched_CatKey,">="&"{cat}|"&TEXT({fy_start},"YYYY-MM"),'
            f'Sched_CatKey,"<="&"{cat}|"&TEXT({rp_end},"YYYY-MM"))'
        )
    if key == "disp_ad":
        return (
            f'=-SUMIFS(Disposal_AD,Disposal_Cat,"{cat}",'
            f'Disposal_Date,">="&({fy_start}),Disposal_Date,"<="&({rp_end}))'
        )
    if key == "close_ad":
        return (
            f'=SUMIFS(Sched_DeprFinal,Sched_CatKey,">="&"{cat}|",'
            f'Sched_CatKey,"<="&"{cat}|"&TEXT({rp_end},"YYYY-MM"))'
            f'+SUMIFS(Register_Opening_AD,Register_Category,"{cat}")'
            f'-SUMIFS(Disposal_AD,Disposal_Cat,"{cat}",Disposal_Date,"<="&({rp_end}))'
        )
    if key == "nbv":
        # Closing Cost - Closing AD, per category. Same lexical CatKey bounds as close_ad.
        return (
            f'=(SUMIFS(Register_Cost,Register_Category,"{cat}",'
            f'Register_InServiceDate,"<="&({rp_end}))'
            f'-SUMIFS(Disposal_Cost,Disposal_Cat,"{cat}",Disposal_Date,"<="&({rp_end})))'
            f'-(SUMIFS(Sched_DeprFinal,Sched_CatKey,">="&"{cat}|",'
            f'Sched_CatKey,"<="&"{cat}|"&TEXT({rp_end},"YYYY-MM"))'
            f'+SUMIFS(Register_Opening_AD,Register_Category,"{cat}")'
            f'-SUMIFS(Disposal_AD,Disposal_Cat,"{cat}",Disposal_Date,"<="&({rp_end})))'
        )
    return ""


# ---------------------------------------------------------------------------
# Tab 6: Disposal Log
# ---------------------------------------------------------------------------

DISP_HEADERS = [
    "Asset ID", "Disposal Date", "Disposal Type", "Commercial Substance? (Trade-in only)",
    "Proceeds", "Authorization (approver name)", "Authorization Date",
    "Evidence Type", "Evidence Reference", "Counterparty",
    "Disposal Cost (calc)", "Accum Depr at Disposal (calc)", "NBV at Disposal (calc)",
    "Gain / (Loss) (calc)", "Notes",
    # Hidden helpers:
    "_Category (calc)",
    "_InPeriodNBV (calc)",
]


def build_disposal_log(wb: Workbook, variant: str):
    ws = wb.create_sheet("Disposal Log")
    ws.sheet_properties.tabColor = "FFF4B084"

    set_col_widths(ws, {
        "A": 12, "B": 14, "C": 16, "D": 18, "E": 14, "F": 22, "G": 14,
        "H": 18, "I": 18, "J": 22, "K": 14, "L": 16, "M": 16, "N": 14, "O": 28, "P": 22,
        "Q": 22,
    })

    for j, h in enumerate(DISP_HEADERS, start=1):
        c = ws.cell(row=1, column=j, value=h)
        style_header(c)
    ws.row_dimensions[1].height = 32

    rows_n = 25 if variant == "paid" else 5
    disposals = DISPOSALS_PAID if variant == "paid" else DISPOSALS_FREE
    reg_end = wb["Asset Register"]._anchors["end_row"]
    sched_end = wb["Schedule"]._anchors["end_row"]

    for i in range(rows_n):
        r = 2 + i
        # Input data
        if i < len(disposals):
            d = disposals[i]
            ws.cell(row=r, column=1, value=d[0])
            ws.cell(row=r, column=2, value=d[1]).number_format = "yyyy-mm-dd"
            ws.cell(row=r, column=3, value=d[2])
            ws.cell(row=r, column=4, value=d[3])
            ws.cell(row=r, column=5, value=d[4]).number_format = "#,##0.00"
            ws.cell(row=r, column=6, value=d[5])
            ws.cell(row=r, column=7, value=d[6]).number_format = "yyyy-mm-dd"
            ws.cell(row=r, column=8, value=d[7])
            ws.cell(row=r, column=9, value=d[8])
            ws.cell(row=r, column=10, value=d[9])
            ws.cell(row=r, column=15, value=d[10])
            for col_idx in range(1, 16):
                ws.cell(row=r, column=col_idx).fill = EXAMPLE_FILL

        # Calc cols K, L, M, N, P
        # K: Disposal Cost = INDEX(Register.L, MATCH(A, Register.A, 0))
        ws.cell(row=r, column=11, value=(
            f'=IF(LEN(A{r})=0,"",IFERROR(INDEX(Register_Cost,MATCH(A{r},Register_AssetID,0)),0))'
        )).number_format = "#,##0.00"
        # L: Accum Depr at Disposal = SUMIFS(Sched_DeprFinal, Sched_AssetID_Col, A, Sched_Period_Col, <=B) + Opening AD
        ws.cell(row=r, column=12, value=(
            f'=IF(LEN(A{r})=0,"",SUMIFS(Sched_DeprFinal,Sched_AssetID_Col,A{r},Sched_Period_Col,"<="&B{r})'
            f'+IFERROR(INDEX(Register_Opening_AD,MATCH(A{r},Register_AssetID,0)),0))'
        )).number_format = "#,##0.00"
        # M: NBV at Disposal = K - L
        ws.cell(row=r, column=13, value=f'=IF(LEN(A{r})=0,"",K{r}-L{r})').number_format = "#,##0.00"
        # N: Gain/(Loss) = E - M
        ws.cell(row=r, column=14, value=f'=IF(LEN(A{r})=0,"",E{r}-M{r})').number_format = "#,##0.00"
        # P: _Category helper
        ws.cell(row=r, column=16, value=(
            f'=IF(LEN(A{r})=0,"",IFERROR(INDEX(Register_Category,MATCH(A{r},Register_AssetID,0)),""))'
        ))
        # Q: _InPeriodNBV helper — NBV-at-disposal for rows whose disposal date is on/before
        # the Reporting Period. Lets Reconciliation Check 4 Side A be period-gated independently
        # of Side B (preserves the cross-check).
        ws.cell(row=r, column=17, value=(
            f'=IF(LEN(A{r})=0,0,IF(B{r}<=Setup!$B$4,M{r},0))'
        )).number_format = "#,##0.00"

        for col_idx in range(1, 18):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = BORDER_ALL
            if col_idx in (11, 12, 13, 14, 16, 17):
                cell.protection = LOCKED
            else:
                cell.protection = UNLOCKED
                if not cell.fill or cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = INPUT_FILL

    # Validations
    end_row = 1 + rows_n
    dv_aid = DataValidation(type="list", formula1=f"=Register_AssetID", allow_blank=True)
    dv_aid.add(f"A2:A{end_row}")
    ws.add_data_validation(dv_aid)

    dv_type = DataValidation(type="list", formula1="=DisposalTypeList", allow_blank=True)
    dv_type.add(f"C2:C{end_row}")
    ws.add_data_validation(dv_type)

    dv_cs = DataValidation(type="list", formula1='"Y,N,N/A"', allow_blank=True)
    dv_cs.add(f"D2:D{end_row}")
    ws.add_data_validation(dv_cs)

    dv_ev = DataValidation(type="list", formula1="=EvidenceTypeList", allow_blank=True)
    dv_ev.add(f"H2:H{end_row}")
    ws.add_data_validation(dv_ev)

    # Hide helper columns P and Q
    ws.column_dimensions["P"].hidden = True
    ws.column_dimensions["Q"].hidden = True

    ws.freeze_panes = "C2"
    ws.protection.set_password(SHEET_PW)
    ws.protection.sheet = True
    ws._anchors = {"end_row": end_row}


# ---------------------------------------------------------------------------
# Tab 7: Change Log
# ---------------------------------------------------------------------------

def build_change_log(wb: Workbook, variant: str):
    ws = wb.create_sheet("Change Log")
    ws.sheet_properties.tabColor = "FFC9C9C9"

    set_col_widths(ws, {
        "A": 12, "B": 22, "C": 24, "D": 24, "E": 16, "F": 48, "G": 22, "H": 22,
    })

    headers = [
        "Asset ID", "Field Changed", "Prior Value", "New Value",
        "Effective Period", "Reason (min 20 chars)", "Approver", "Timestamp",
    ]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        style_header(c)
    ws.row_dimensions[1].height = 32

    rows_n = 50 if variant == "paid" else 10
    for i in range(rows_n):
        r = 2 + i
        for col_idx in range(1, 9):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = BORDER_ALL
            cell.protection = UNLOCKED
            cell.fill = INPUT_FILL
        ws.cell(row=r, column=5).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=8).number_format = "yyyy-mm-dd hh:mm"

    end_row = 1 + rows_n
    dv_aid = DataValidation(type="list", formula1="=Register_AssetID", allow_blank=True)
    dv_aid.add(f"A2:A{end_row}")
    ws.add_data_validation(dv_aid)

    dv_field = DataValidation(
        type="list",
        formula1='"Useful Life,Method,Opening Accum Depr,Category"',
        allow_blank=True,
    )
    dv_field.add(f"B2:B{end_row}")
    ws.add_data_validation(dv_field)

    # Reason min 20 chars
    dv_reason = DataValidation(
        type="custom",
        formula1=f"=OR(LEN(F2)=0,LEN(F2)>=20)",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Reason too short",
        error="Reason must be at least 20 characters (ASC 250 disclosure).",
    )
    dv_reason.add(f"F2:F{end_row}")
    ws.add_data_validation(dv_reason)

    ws.protection.set_password(SHEET_PW)
    ws.protection.sheet = True


# ---------------------------------------------------------------------------
# Tab 8: Reconciliation + Diagnostic
# ---------------------------------------------------------------------------

def build_reconciliation(wb: Workbook, variant: str):
    ws = wb.create_sheet("Reconciliation")
    ws.sheet_properties.tabColor = "FFE2EFDA"

    set_col_widths(ws, {"A": 4, "B": 44, "C": 18, "D": 18, "E": 18, "F": 22, "G": 60})

    ws["B1"] = "Five-Way Reconciliation — Reporting Period"
    ws["B1"].font = TITLE_FONT
    ws["F1"] = "=Setup!$B$4"
    ws["F1"].number_format = "yyyy-mm-dd"

    headers = ["Check", "Side A", "Side B", "Variance (A − B)", "Status"]
    for j, h in enumerate(headers, start=2):
        c = ws.cell(row=3, column=j, value=h)
        style_header(c)

    # 5 checks
    # Check 1: Schedule Depr Total (this month) = JE Depr Total (sum of category DR rows on JE Generator)
    # Check 2: Schedule Accum Total (running through RP) = Rollforward Closing Accum TOTAL
    # Check 3: Register Cost = Rollforward Cost Closing TOTAL
    # Check 4: Register Disposal Adjustment = Disposal Log net (NBV)
    # Check 5: NBV Math: Cost Closing - Accum Closing - Rollforward NBV TOTAL = 0
    rp_end = "Setup!$B$4"
    fy_start = "Setup!$B$3"
    base = 4

    # Helper: Rollforward references — we need to read total column of specific rows.
    # In Rollforward we set base_r=4, so:
    #   row 4: opening_cost
    #   row 5: add_cost
    #   row 6: disp_cost
    #   row 7: close_cost
    #   row 8: opening_ad
    #   row 9: depr_exp
    #   row 10: disp_ad
    #   row 11: close_ad
    #   row 12: nbv
    # TOTAL is in column 2 + len(CATEGORIES) = column 8 (H)
    total_col = get_column_letter(2 + len(CATEGORIES))
    rf = f"Rollforward!${total_col}$"
    rf_close_cost = f"{rf}7"
    rf_depr_exp = f"{rf}9"
    rf_close_ad = f"{rf}11"
    rf_nbv = f"{rf}12"

    checks = [
        {
            "name": "Schedule Depr (period) = JE Generator Depr",
            "side_a": f'=SUMIFS(Sched_DeprFinal,Sched_Period_Col,{rp_end})',
            "side_b": f'=SUMIFS(\'JE Generator\'!D5:D{4 + len(CATEGORIES)},'
                      f'\'JE Generator\'!A5:A{4 + len(CATEGORIES)},{rp_end})',
        },
        {
            "name": "Schedule Accum Closing = Rollforward Closing Accum TOTAL",
            "side_a": (
                f'=SUMIFS(Sched_DeprFinal,Sched_Period_Col,"<="&{rp_end})'
                f'+SUM(Register_Opening_AD)'
                f'-SUMIFS(Disposal_AD,Disposal_Date,"<="&{rp_end})'
            ),
            "side_b": f"={rf_close_ad}",
        },
        {
            "name": "Register Cost (not disposed) = Rollforward Cost Closing TOTAL",
            "side_a": (
                f'=SUMIFS(Register_Cost,Register_InServiceDate,"<="&{rp_end})'
                f'-SUMIFS(Disposal_Cost,Disposal_Date,"<="&{rp_end})'
            ),
            "side_b": f"={rf_close_cost}",
        },
        {
            "name": "Disposal NBV total (in-period) = Disposal Log Σ(NBV at Disposal)",
            "side_a": f'=SUM(\'Disposal Log\'!Q2:Q{wb["Disposal Log"]._anchors["end_row"]})',
            "side_b": (
                f'=SUMIFS(Disposal_Cost,Disposal_Date,"<="&{rp_end})'
                f'-SUMIFS(Disposal_AD,Disposal_Date,"<="&{rp_end})'
            ),
        },
        {
            "name": "NBV math: Cost Closing − Accum Closing = Rollforward NBV TOTAL",
            "side_a": f"={rf_close_cost}-{rf_close_ad}",
            "side_b": f"={rf_nbv}",
        },
    ]

    for i, ch in enumerate(checks):
        r = base + i
        ws.cell(row=r, column=2, value=ch["name"])
        ws.cell(row=r, column=3, value=ch["side_a"]).number_format = "#,##0.00"
        ws.cell(row=r, column=4, value=ch["side_b"]).number_format = "#,##0.00"
        ws.cell(row=r, column=5, value=f"=C{r}-D{r}").number_format = "#,##0.00"
        # Status text mirroring CF
        ws.cell(row=r, column=6, value=(
            f'=IF(ABS(E{r})<0.01,"TIES",IF(ABS(E{r})<=1,"Rounding","VARIANCE"))'
        ))
        for c_idx in range(2, 7):
            ws.cell(row=r, column=c_idx).border = BORDER_ALL
            ws.cell(row=r, column=c_idx).protection = LOCKED

    # Conditional formatting on variance E column
    cf_range = f"E{base}:E{base + len(checks) - 1}"
    ws.conditional_formatting.add(cf_range, FormulaRule(
        formula=[f"ABS(E{base})<0.01"], fill=GREEN_FILL,
    ))
    ws.conditional_formatting.add(cf_range, FormulaRule(
        formula=[f"AND(ABS(E{base})>=0.01,ABS(E{base})<=1)"], fill=YELLOW_FILL,
    ))
    ws.conditional_formatting.add(cf_range, FormulaRule(
        formula=[f"ABS(E{base})>1"], fill=RED_FILL,
    ))

    # Diagnostic sub-section
    diag_r = base + len(checks) + 2
    ws.cell(row=diag_r, column=2, value="Diagnostic — likely causes when red")
    style_section(ws.cell(row=diag_r, column=2))

    diagnostics = [
        ("Disposal logged but Register row not updated — check Status",
         "'Asset Register'!A2", "Jump to Register first row"),
        ("Asset acquired after Reporting Period close — check Acquisition Date",
         "'Asset Register'!J2", "Jump to Register Acquisition Date"),
        ("Opening Accum Depr entered without Opening Period — check S/T",
         "'Asset Register'!S2", "Jump to Register Opening Accum Depr"),
        ("Useful Life or Method revised but no Change Log entry — verify ASC 250 trail",
         "'Change Log'!A2", "Jump to Change Log"),
        ("Edit detected in a Closed Period — review Setup Closed Periods table",
         "Setup!A1", "Jump to Setup Closed Periods"),
        ("New Category doesn't match a defined Policy row — review Policy table",
         "Setup!A1", "Jump to Setup Policy"),
    ]
    for i, (label, target, link_text) in enumerate(diagnostics):
        r = diag_r + 1 + i
        ws.cell(row=r, column=2, value=f"{i + 1}. {label}")
        ws.cell(row=r, column=3, value=f'=HYPERLINK("#{target}","{link_text}")')
        ws.cell(row=r, column=3).font = Font(color="FF0563C1", underline="single")
        ws.cell(row=r, column=2).border = BORDER_ALL
        ws.cell(row=r, column=3).border = BORDER_ALL

    ws.freeze_panes = "B4"
    ws.protection.set_password(SHEET_PW)
    ws.protection.sheet = True


# ---------------------------------------------------------------------------
# Tab 9: Audit Confirmation
# ---------------------------------------------------------------------------

def build_audit_confirmation(wb: Workbook, variant: str):
    ws = wb.create_sheet("Audit Confirmation")
    ws.sheet_properties.tabColor = "FFB4C7E7"

    set_col_widths(ws, {
        "A": 12, "B": 30, "C": 22, "D": 14, "E": 14, "F": 10, "G": 14, "H": 14, "I": 14,
    })

    ws["A1"] = "=Setup!$B$2"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Fixed Asset Audit Confirmation — PBC List"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A3"] = "Reporting Period (Month End):"
    ws["B3"] = "=Setup!$B$4"
    ws["B3"].number_format = "yyyy-mm-dd"

    # Sample input
    ws["A5"] = "Auditor Reference:"
    ws["B5"] = ""  # input
    style_input(ws["B5"])
    ws["A6"] = "Sample Size Requested:"
    ws["B6"] = 25
    style_input(ws["B6"])

    # Header row
    headers = ["Asset ID", "Description", "Category", "In-Service Date", "Cost",
               "Useful Life (mo)", "Method", "Accum Depr @ RP", "NBV @ RP"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=8, column=j, value=h)
        style_header(c)
    ws.row_dimensions[8].height = 32

    rp_end = "Setup!$B$4"
    rows_n = 25 if variant == "paid" else 5
    reg_end = wb["Asset Register"]._anchors["end_row"]
    for i in range(rows_n):
        r = 9 + i
        reg_row = 2 + i  # first N assets are the sample
        ws.cell(row=r, column=1, value=f"='Asset Register'!A{reg_row}")
        ws.cell(row=r, column=2, value=f"='Asset Register'!B{reg_row}")
        ws.cell(row=r, column=3, value=f"='Asset Register'!C{reg_row}")
        ws.cell(row=r, column=4, value=f"='Asset Register'!K{reg_row}").number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=5, value=f"='Asset Register'!L{reg_row}").number_format = "#,##0.00"
        ws.cell(row=r, column=6, value=f"='Asset Register'!N{reg_row}")
        ws.cell(row=r, column=7, value=f"='Asset Register'!O{reg_row}")
        # Accum Depr @ RP = SUMIFS on Sched + Opening
        ws.cell(row=r, column=8, value=(
            f'=IF(LEN(\'Asset Register\'!A{reg_row})=0,"",'
            f'SUMIFS(Sched_DeprFinal,Sched_AssetID_Col,\'Asset Register\'!A{reg_row},'
            f'Sched_Period_Col,"<="&{rp_end})+\'Asset Register\'!S{reg_row})'
        )).number_format = "#,##0.00"
        ws.cell(row=r, column=9, value=(
            f'=IF(LEN(\'Asset Register\'!A{reg_row})=0,"",\'Asset Register\'!L{reg_row}-H{r})'
        )).number_format = "#,##0.00"
        for col_idx in range(1, 10):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = BORDER_ALL
            cell.protection = LOCKED

    sig_r = 9 + rows_n + 2
    ws.cell(row=sig_r, column=1, value="Management Representation").font = SECTION_FONT
    ws.cell(row=sig_r + 1, column=1, value=(
        "Management represents that the Fixed Asset Register and Schedule above accurately reflect the"
        " company's fixed assets at the reporting date, in accordance with US GAAP (ASC 360), and that"
        " supporting documentation (invoices, POs, capitalization approvals, evidence of disposal) is"
        " available upon request."
    ))
    ws.cell(row=sig_r + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[sig_r + 1].height = 60
    ws.merge_cells(start_row=sig_r + 1, start_column=1, end_row=sig_r + 1, end_column=9)

    sig_r2 = sig_r + 3
    ws.cell(row=sig_r2, column=1, value="Prepared By:")
    ws.cell(row=sig_r2, column=2, value="")
    style_input(ws.cell(row=sig_r2, column=2))
    ws.cell(row=sig_r2, column=4, value="Date:")
    ws.cell(row=sig_r2, column=5, value="")
    style_input(ws.cell(row=sig_r2, column=5))

    ws.cell(row=sig_r2 + 1, column=1, value="Reviewed By:")
    ws.cell(row=sig_r2 + 1, column=2, value="")
    style_input(ws.cell(row=sig_r2 + 1, column=2))
    ws.cell(row=sig_r2 + 1, column=4, value="Date:")
    ws.cell(row=sig_r2 + 1, column=5, value="")
    style_input(ws.cell(row=sig_r2 + 1, column=5))

    # Page setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
    ws.print_options = PrintOptions(horizontalCentered=True)
    ws.oddHeader.center.text = "&\"-,Bold\"&14 Fixed Asset Audit Confirmation"
    ws.oddFooter.left.text = "&A"
    ws.oddFooter.center.text = "Confidential — for auditor use"
    ws.oddFooter.right.text = "Page &P of &N"
    ws.print_title_rows = "8:8"
    ws.print_area = f"A1:I{sig_r2 + 2}"

    ws.protection.set_password(SHEET_PW)
    ws.protection.sheet = True


# ---------------------------------------------------------------------------
# Tab 10: _Pivot (hidden, long-form pre-pivot table)
# ---------------------------------------------------------------------------

def build_pivot(wb: Workbook, variant: str):
    ws = wb.create_sheet("_Pivot")
    ws.sheet_properties.tabColor = "FFCCCCCC"

    set_col_widths(ws, {"A": 12, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 24})

    headers = ["Asset ID", "Category", "Method", "Period (YYYY-MM)", "Depr This Period", "Status", "CategoryKey"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        style_header(c)
    ws.row_dimensions[1].height = 32

    # Build long-form rows mirroring Schedule
    n_rows = wb["Schedule"]._anchors["total_rows"]
    for i in range(n_rows):
        r = 2 + i
        sr = 2 + i  # Schedule row
        ws.cell(row=r, column=1, value=f"=Schedule!A{sr}")
        # Category by lookup on Register
        ws.cell(row=r, column=2, value=(
            f'=IF(LEN(Schedule!A{sr})=0,"",IFERROR(INDEX(Register_Category,MATCH(Schedule!A{sr},Register_AssetID,0)),""))'
        ))
        ws.cell(row=r, column=3, value=f"=Schedule!E{sr}")
        ws.cell(row=r, column=4, value=f'=IF(LEN(Schedule!A{sr})=0,"",TEXT(Schedule!B{sr},"YYYY-MM"))')
        ws.cell(row=r, column=5, value=f"=Schedule!H{sr}").number_format = "#,##0.00"
        # Status pulled from Register
        ws.cell(row=r, column=6, value=(
            f'=IF(LEN(Schedule!A{sr})=0,"",IFERROR(INDEX(Register_Status,MATCH(Schedule!A{sr},Register_AssetID,0)),""))'
        ))
        ws.cell(row=r, column=7, value=f"=Schedule!L{sr}")

        for c_idx in range(1, 8):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = BORDER_ALL
            cell.protection = LOCKED

    # Note about pivot
    ws.cell(row=1, column=9, value=(
        "PIVOT-READY: Select A1:G{n} and Insert > PivotTable. Rows=Category, Columns=Period, Values=SUM(Depr), Filters=Method/Status.".format(n=1 + n_rows)
    ))
    ws.cell(row=1, column=9).font = SMALL_FONT

    ws.sheet_state = "hidden"
    ws.protection.set_password(SHEET_PW)
    ws.protection.sheet = True


# ---------------------------------------------------------------------------
# Named ranges
# ---------------------------------------------------------------------------

def add_named_ranges(wb: Workbook, variant: str):
    """Create all 32 named ranges scoped to workbook."""
    setup = wb["Setup"]
    s_anc = setup._anchors
    reg = wb["Asset Register"]
    reg_end = reg._anchors["end_row"]
    sched = wb["Schedule"]
    sched_end = sched._anchors["end_row"]
    disp = wb["Disposal Log"]
    disp_end = disp._anchors["end_row"]

    def addn(name: str, ref: str):
        dn = DefinedName(name=name, attr_text=ref)
        wb.defined_names[name] = dn

    # Setup single-cell refs
    addn("Co_Name", f"Setup!${s_anc['Co_Name'][:1]}${s_anc['Co_Name'][1:]}")
    addn("FY_Start", f"Setup!${s_anc['FY_Start'][:1]}${s_anc['FY_Start'][1:]}")
    addn("ReportingPeriodEnd", f"Setup!${s_anc['ReportingPeriodEnd'][:1]}${s_anc['ReportingPeriodEnd'][1:]}")
    addn("DefaultConvention", f"Setup!${s_anc['DefaultConvention'][:1]}${s_anc['DefaultConvention'][1:]}")
    addn("DefaultMethod", f"Setup!${s_anc['DefaultMethod'][:1]}${s_anc['DefaultMethod'][1:]}")
    addn("Materiality", f"Setup!${s_anc['Materiality'][:1]}${s_anc['Materiality'][1:]}")

    # GL codes
    for nm in ("GL_ComputerEq", "GL_Furniture", "GL_LHI", "GL_Machinery",
               "GL_Vehicles", "GL_Other", "GL_AccumDepr", "GL_DeprExpense",
               "GL_DispGain", "GL_DispLoss"):
        ref = s_anc[nm]
        addn(nm, f"Setup!${ref[:1]}${ref[1:]}")

    # Lists
    def range_str(start_col, start_row, end_col, end_row):
        return f"Setup!${get_column_letter(start_col)}${start_row}:${get_column_letter(end_col)}${end_row}"

    addn("CategoryList", range_str(*s_anc["CategoryList_Range"]))
    addn("MethodList", range_str(*s_anc["MethodList_Range"]))
    addn("ConventionList", range_str(*s_anc["ConventionList_Range"]))
    addn("DisposalTypeList", range_str(*s_anc["DisposalTypeList_Range"]))
    addn("EvidenceTypeList", range_str(*s_anc["EvidenceTypeList_Range"]))
    if variant == "paid":
        addn("GLSystemList", range_str(*s_anc["GLSystemList_Range"]))

    # Policy table — separate columns
    p_start = s_anc["Policy_Cat_Start_Row"]
    p_end = s_anc["Policy_Cat_End_Row"]
    addn("Policy_Category", f"Setup!$A${p_start}:$A${p_end}")
    addn("Policy_Life", f"Setup!$B${p_start}:$B${p_end}")
    addn("Policy_Method", f"Setup!$C${p_start}:$C${p_end}")
    addn("Policy_Salvage", f"Setup!$D${p_start}:$D${p_end}")

    # Closed Periods
    cp_start = s_anc["ClosedPeriods_Start_Row"]
    cp_end = s_anc["ClosedPeriods_End_Row"]
    addn("ClosedPeriods", f"Setup!$A${cp_start}:$A${cp_end}")

    # Register columns
    addn("Register_AssetID",       f"'Asset Register'!$A$2:$A${reg_end}")
    addn("Register_Cost",          f"'Asset Register'!$L$2:$L${reg_end}")
    addn("Register_Salvage",       f"'Asset Register'!$M$2:$M${reg_end}")
    addn("Register_Life",          f"'Asset Register'!$N$2:$N${reg_end}")
    addn("Register_Method",        f"'Asset Register'!$O$2:$O${reg_end}")
    addn("Register_Category",      f"'Asset Register'!$C$2:$C${reg_end}")
    addn("Register_InServiceDate", f"'Asset Register'!$K$2:$K${reg_end}")
    addn("Register_Opening_AD",    f"'Asset Register'!$S$2:$S${reg_end}")
    addn("Register_Status",        f"'Asset Register'!$AC$2:$AC${reg_end}")

    # Schedule columns
    addn("Sched_AssetID_Col", f"Schedule!$A$2:$A${sched_end}")
    addn("Sched_AssetID",     f"Schedule!$A$2:$A${sched_end}")
    addn("Sched_Period_Col",  f"Schedule!$B$2:$B${sched_end}")
    addn("Sched_Period",      f"Schedule!$B$2:$B${sched_end}")
    addn("Sched_DeprFinal",   f"Schedule!$H$2:$H${sched_end}")
    addn("Sched_Depr",        f"Schedule!$H$2:$H${sched_end}")
    addn("Sched_AccumDepr",   f"Schedule!$I$2:$I${sched_end}")
    addn("Sched_CatKey",      f"Schedule!$L$2:$L${sched_end}")

    # Disposal Log
    addn("Disposal_AssetID", f"'Disposal Log'!$A$2:$A${disp_end}")
    addn("Disp_AssetID",     f"'Disposal Log'!$A$2:$A${disp_end}")
    addn("Disposal_Date",    f"'Disposal Log'!$B$2:$B${disp_end}")
    addn("Disp_Date",        f"'Disposal Log'!$B$2:$B${disp_end}")
    addn("Disposal_Proceeds", f"'Disposal Log'!$E$2:$E${disp_end}")
    addn("Disp_Proceeds",    f"'Disposal Log'!$E$2:$E${disp_end}")
    addn("Disposal_Cost",    f"'Disposal Log'!$K$2:$K${disp_end}")
    addn("Disposal_AD",      f"'Disposal Log'!$L$2:$L${disp_end}")
    addn("Disposal_Cat",     f"'Disposal Log'!$P$2:$P${disp_end}")


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def build_workbook(variant: str) -> pathlib.Path:
    assert variant in ("paid", "free"), variant
    wb = Workbook()
    # Remove default
    default = wb.active
    wb.remove(default)

    # Build in dependency order
    build_readme(wb, variant)            # 0
    build_setup(wb, variant)             # 1
    build_register(wb, variant)          # 2
    build_schedule(wb, variant)          # 3
    build_disposal_log(wb, variant)      # need register anchors; built before JE/Roll/Recon/Audit
    build_je_generator(wb, variant)      # 4
    build_rollforward(wb, variant)       # 5
    build_change_log(wb, variant)        # 7
    build_reconciliation(wb, variant)    # 8
    build_audit_confirmation(wb, variant) # 9
    build_pivot(wb, variant)             # 10 hidden

    # Reorder sheets to spec order: README, Setup, Asset Register, Schedule,
    # JE Generator, Rollforward, Disposal Log, Change Log, Reconciliation,
    # Audit Confirmation, _Pivot
    desired = [
        "README", "Setup", "Asset Register", "Schedule", "JE Generator",
        "Rollforward", "Disposal Log", "Change Log", "Reconciliation",
        "Audit Confirmation", "_Pivot",
    ]
    if variant == "free":
        # No _Pivot in free? Spec says free is "otherwise identical structure"; keep _Pivot.
        pass
    # openpyxl: sheets order via wb._sheets list
    new_order = []
    for name in desired:
        if name in wb.sheetnames:
            new_order.append(wb[name])
    other = [s for s in wb._sheets if s not in new_order]
    wb._sheets = new_order + other

    # Named ranges
    add_named_ranges(wb, variant)

    # Workbook-level metadata
    wb.properties.title = "KDesk Accounting — Fixed Asset Rollforward Workbook"
    wb.properties.creator = "KDesk Accounting"
    wb.properties.subject = "Fixed Asset Rollforward (ASC 360 book depreciation)"
    wb.properties.keywords = "fixed assets, ASC 360, rollforward, depreciation, audit-ready"
    wb.properties.description = (
        "Audit-ready fixed asset rollforward workbook with policy table, "
        "change log, audit confirmation export, and 5-way reconciliation. "
        "Built by KDesk Accounting — kdeskaccounting.com"
    )

    suffix = "v1" if variant == "paid" else "free-v1"
    out_path = DIST / f"fixed-asset-rollforward-{suffix}.xlsx"
    DIST.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--variant", choices=("paid", "free", "both"), default="both")
    args = ap.parse_args()

    targets = ("paid", "free") if args.variant == "both" else (args.variant,)
    for v in targets:
        path = build_workbook(v)
        size_kb = path.stat().st_size / 1024
        print(f"Built {v}: {path}  ({size_kb:,.1f} KB)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
