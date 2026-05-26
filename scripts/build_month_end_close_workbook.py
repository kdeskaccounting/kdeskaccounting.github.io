#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "openpyxl>=3.1",
# ]
# ///
"""
Build the KDesk Accounting Month-End Close Checklist + Tie-Out Workbook.

Free lead-magnet companion to the paid Fixed Asset / ASC 606 / ASC 842 workbooks.

Produces:
- dist/templates/month-end-close-checklist-v1.xlsx

Five tabs:
  1. README + Setup       Company config, materiality, quickstart, upsell
  2. Close Calendar       ~40 pre-populated tasks across 5 phases (Day 1-2 .. Day 5+)
  3. Reconciliations      18 standard subledger-to-GL ties with variance + flag
  4. JE Tracker           Period JE log with debit=credit sanity check
  5. Sign-Off             Printable certification page summarizing the close

Function library: Excel 2016/365/Mac safe set (no XLOOKUP, LET, LAMBDA, IFS,
dynamic-array spills). All formulas tested via openpyxl re-load smoke test.
"""
from __future__ import annotations

import argparse
import datetime as dt
import pathlib
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

REPO_ROOT = pathlib.Path(__file__).resolve().parents[1]
DIST = REPO_ROOT / "dist" / "templates"

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

COLOR_NAVY = "FF1F4E78"
COLOR_WHITE = "FFFFFFFF"
COLOR_SECTION_BG = "FFD9E1F2"
COLOR_INPUT_BG = "FFFFF2CC"
COLOR_BORDER = "FFB7B7B7"
COLOR_GREEN = "FFC6EFCE"
COLOR_GREEN_FG = "FF006100"
COLOR_YELLOW = "FFFFEB9C"
COLOR_YELLOW_FG = "FF9C5700"
COLOR_RED = "FFFFC7CE"
COLOR_RED_FG = "FF9C0006"
COLOR_GRAY = "FFE7E6E6"
COLOR_GRAY_FG = "FF595959"

THIN = Side(style="thin", color=COLOR_BORDER)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_HEADER = PatternFill("solid", fgColor=COLOR_NAVY)
FILL_SECTION = PatternFill("solid", fgColor=COLOR_SECTION_BG)
FILL_INPUT = PatternFill("solid", fgColor=COLOR_INPUT_BG)

FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=COLOR_NAVY)
FONT_H2 = Font(name="Calibri", size=13, bold=True, color=COLOR_NAVY)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=COLOR_WHITE)
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_NOTE = Font(name="Calibri", size=10, italic=True, color=COLOR_GRAY_FG)
FONT_LINK = Font(name="Calibri", size=11, color="FF2E75B6", underline="single")

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# ---------------------------------------------------------------------------
# Domain constants
# ---------------------------------------------------------------------------

STATUS_OPTIONS = ["Not Started", "In Progress", "Done", "Blocked", "N/A"]
JE_SOURCE_OPTIONS = [
    "Manual",
    "FA Workbook (KDesk)",
    "ASC 606 Workbook (KDesk)",
    "ASC 842 Workbook (KDesk)",
    "GL System Auto",
    "Other",
]
YES_NO = ["Yes", "No", "N/A"]

# Pre-populated tasks: (phase, task, owner_default)
TASKS: list[tuple[str, str, str]] = [
    # Phase 1: Day 1-2 cutoff
    ("Day 1-2 — Cutoff & Data Collection", "Confirm all revenue recognized in the period has a corresponding invoice or contract milestone", "Revenue Accountant"),
    ("Day 1-2 — Cutoff & Data Collection", "Identify deals signed in the last 3 days — confirm period recognition", "Revenue Accountant"),
    ("Day 1-2 — Cutoff & Data Collection", "Pull deferred revenue balance from subledger and reconcile to GL", "Revenue Accountant"),
    ("Day 1-2 — Cutoff & Data Collection", "Post all vendor invoices received through last business day", "AP Accountant"),
    ("Day 1-2 — Cutoff & Data Collection", "Accrue for goods/services received but not yet invoiced (legal, consulting, cloud)", "AP Accountant"),
    ("Day 1-2 — Cutoff & Data Collection", "Confirm purchase order cutoff with operations / procurement", "AP Accountant"),
    ("Day 1-2 — Cutoff & Data Collection", "Accrue wages earned but not paid (partial pay periods)", "Payroll"),
    ("Day 1-2 — Cutoff & Data Collection", "Post payroll tax and benefits accruals", "Payroll"),
    ("Day 1-2 — Cutoff & Data Collection", "Update commission expense accruals (see ASC 606 schedule)", "Payroll / Comp"),
    ("Day 1-2 — Cutoff & Data Collection", "Pull final bank statements for the last day of the period", "Treasury"),
    ("Day 1-2 — Cutoff & Data Collection", "Post unrecorded bank items (fees, interest, wires)", "Treasury"),
    ("Day 1-2 — Cutoff & Data Collection", "Confirm all credit card charges imported and coded", "AP Accountant"),

    # Phase 2: Day 2-3 subledger
    ("Day 2-3 — Subledger Reconciliations", "AR subledger balance ties to GL", "AR Accountant"),
    ("Day 2-3 — Subledger Reconciliations", "Aged AR report reviewed — flag items >90 days for reserve analysis", "AR Accountant"),
    ("Day 2-3 — Subledger Reconciliations", "Bad debt reserve updated (if allowance method)", "AR Accountant"),
    ("Day 2-3 — Subledger Reconciliations", "Credit memos and unapplied cash resolved", "AR Accountant"),
    ("Day 2-3 — Subledger Reconciliations", "AP subledger balance ties to GL", "AP Accountant"),
    ("Day 2-3 — Subledger Reconciliations", "Vendor statements reconciled for key vendors", "AP Accountant"),
    ("Day 2-3 — Subledger Reconciliations", "Outstanding checks reviewed for stale items (>90 days)", "AP Accountant"),
    ("Day 2-3 — Subledger Reconciliations", "Fixed asset additions and disposals posted for the period", "Fixed Asset Accountant"),
    ("Day 2-3 — Subledger Reconciliations", "Depreciation run for all assets (see FA Rollforward Workbook)", "Fixed Asset Accountant"),
    ("Day 2-3 — Subledger Reconciliations", "Accumulated depreciation ties to fixed asset subledger", "Fixed Asset Accountant"),

    # Phase 3: Day 3-4 technical schedules
    ("Day 3-4 — Technical Accounting Schedules", "Deferred commission new payments entered (see ASC 606 Workbook)", "Revenue Accountant"),
    ("Day 3-4 — Technical Accounting Schedules", "Period amortization expense calculated and JE posted", "Revenue Accountant"),
    ("Day 3-4 — Technical Accounting Schedules", "Deferred commission asset balance per schedule ties to GL", "Revenue Accountant"),
    ("Day 3-4 — Technical Accounting Schedules", "Monthly lease JE generated per active lease (see ASC 842 Workbook)", "Technical Accountant"),
    ("Day 3-4 — Technical Accounting Schedules", "Lease liability balance per amortization schedule ties to GL", "Technical Accountant"),
    ("Day 3-4 — Technical Accounting Schedules", "ROU asset balance per schedule ties to GL", "Technical Accountant"),
    ("Day 3-4 — Technical Accounting Schedules", "Any new leases commenced — initial recognition entry posted", "Technical Accountant"),
    ("Day 3-4 — Technical Accounting Schedules", "Prepaid schedule updated with new payments; monthly amortization posted", "Staff Accountant"),
    ("Day 3-4 — Technical Accounting Schedules", "Recurring accruals posted (insurance, rent, professional fees)", "Staff Accountant"),
    ("Day 3-4 — Technical Accounting Schedules", "Prior-period one-time accruals reversed if appropriate", "Staff Accountant"),

    # Phase 4: Day 4-5 equity / tax / other
    ("Day 4-5 — Equity, Tax & Other", "Stock-based comp expense calculated and posted (RSUs, options)", "Technical Accountant"),
    ("Day 4-5 — Equity, Tax & Other", "Tax provision estimate adjusted; deferred tax true-up if needed", "Tax / Controller"),
    ("Day 4-5 — Equity, Tax & Other", "FX revaluation posted (if foreign-currency balances)", "Technical Accountant"),
    ("Day 4-5 — Equity, Tax & Other", "Inter-company eliminations posted (if multi-entity)", "Technical Accountant"),

    # Phase 5: Day 5+ statements & review
    ("Day 5+ — Statements & Review", "Trial balance pulled; debits = credits", "Controller"),
    ("Day 5+ — Statements & Review", "Balance sheet, income statement, and cash flow generated", "Controller"),
    ("Day 5+ — Statements & Review", "Flux / variance analysis vs prior period (significant items explained)", "Controller"),
    ("Day 5+ — Statements & Review", "Management review with CFO / CEO; adjustments posted", "Controller / CFO"),
    ("Day 5+ — Statements & Review", "Period locked in GL system", "Controller"),
    ("Day 5+ — Statements & Review", "Close package distributed to leadership / board", "Controller / CFO"),
]

# Pre-populated reconciliations: (account_name, source/schedule, category)
RECONCILIATIONS: list[tuple[str, str]] = [
    ("Cash — Operating Account", "Bank statement / Treasury workbook"),
    ("Cash — Money Market / Savings", "Bank statement"),
    ("Accounts Receivable", "AR aging subledger"),
    ("Allowance for Doubtful Accounts", "Reserve calculation"),
    ("Prepaid Expenses", "Prepaid amortization schedule"),
    ("Deferred Commissions (ASC 606)", "ASC 606 Commission Workbook"),
    ("Fixed Assets — Cost", "Fixed Asset Register"),
    ("Accumulated Depreciation", "FA Rollforward Workbook"),
    ("ROU Asset — Operating Leases", "ASC 842 Lease Workbook"),
    ("ROU Asset — Finance Leases", "ASC 842 Lease Workbook"),
    ("Accounts Payable", "AP aging subledger"),
    ("Accrued Liabilities", "Accrual schedule"),
    ("Accrued Payroll & Benefits", "Payroll provider report"),
    ("Lease Liability — Operating", "ASC 842 Lease Workbook"),
    ("Lease Liability — Finance", "ASC 842 Lease Workbook"),
    ("Deferred Revenue", "Revenue recognition subledger"),
    ("Common Stock & APIC", "Cap table / equity ledger"),
    ("Retained Earnings", "Roll-forward from prior period"),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def set_cell(ws, ref, value, font=None, fill=None, align=None, border=None, number_format=None):
    c = ws[ref]
    c.value = value
    if font is not None:
        c.font = font
    if fill is not None:
        c.fill = fill
    if align is not None:
        c.alignment = align
    if border is not None:
        c.border = border
    if number_format is not None:
        c.number_format = number_format
    return c


def write_table_header(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        col = get_column_letter(start_col + i)
        set_cell(ws, f"{col}{row}", h, font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_CENTER, border=BORDER_ALL)
    ws.row_dimensions[row].height = 32


def set_col_widths(ws, widths: dict[str, float]):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def add_defined_name(wb, name, ref):
    """Workbook-level defined name. ref is like 'Setup!$B$5'."""
    dn = DefinedName(name=name, attr_text=ref)
    wb.defined_names[name] = dn


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_readme_setup(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README & Setup"
    set_col_widths(ws, {"A": 4, "B": 28, "C": 50, "D": 24, "E": 24, "F": 24})
    ws.sheet_view.showGridLines = False

    # Title band
    ws.merge_cells("B2:F2")
    set_cell(ws, "B2", "Month-End Close Checklist + Tie-Out Workbook",
             font=Font(name="Calibri", size=22, bold=True, color=COLOR_WHITE),
             fill=FILL_HEADER, align=Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[2].height = 42

    ws.merge_cells("B3:F3")
    set_cell(ws, "B3", "by KDesk Accounting — free companion to our paid workbooks. v1.0",
             font=FONT_NOTE, align=Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[3].height = 20

    # Quickstart
    set_cell(ws, "B5", "Quickstart", font=FONT_H2)
    quickstart = [
        "1. Fill in your company info in the Setup table below (yellow cells).",
        "2. Go to the Close Calendar tab. Review the 42 pre-populated tasks. Adjust Owner and Due Date for your team. Add or delete rows as needed.",
        "3. Work through tasks during the close. Set Status as you go — conditional formatting flags Blocked items in red.",
        "4. On Reconciliations, enter GL Balance and Subledger Balance for each account. The Variance and Flag columns auto-calculate against the Materiality threshold from Setup.",
        "5. Log every JE posted on JE Tracker. The Debits = Credits sanity check at the bottom catches one-sided entries.",
        "6. Finish on Sign-Off. The summary metrics auto-populate. Have Preparer, Reviewer, and Controller sign off, lock the period in your GL, and distribute the close package.",
    ]
    for i, line in enumerate(quickstart):
        set_cell(ws, f"B{6+i}", line, font=FONT_BODY, align=ALIGN_LEFT_TOP)
        ws.merge_cells(f"B{6+i}:F{6+i}")
        ws.row_dimensions[6+i].height = 28

    # Setup table
    setup_row = 14
    set_cell(ws, f"B{setup_row}", "Setup", font=FONT_H2)
    setup_fields = [
        ("Company Name", "Acme Corp.", "text"),
        ("Fiscal Year End (MM-DD)", "12-31", "text"),
        ("Current Period (e.g., May 2026)", "May 2026", "text"),
        ("Period End Date", dt.date(2026, 5, 31), "date"),
        ("Performance Materiality ($)", 50000, "currency"),
        ("Reporting Currency", "USD", "text"),
        ("Period Locked in GL?", "No", "yesno"),
    ]
    write_table_header(ws, setup_row + 1, ["#", "Field", "Value"], start_col=2)
    # Wait: write_table_header uses col widths starting at B (col 2). We want #, Field, Value across B, C, D.
    # Actually let me redo this — use B (label) D (value) layout, simpler.
    # Clear and redo:
    set_cell(ws, f"B{setup_row+1}", "Field", font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_CENTER, border=BORDER_ALL)
    set_cell(ws, f"C{setup_row+1}", "Value", font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_CENTER, border=BORDER_ALL)
    ws.merge_cells(f"C{setup_row+1}:D{setup_row+1}")
    ws.row_dimensions[setup_row+1].height = 28

    value_cells = {}
    for i, (label, default, kind) in enumerate(setup_fields):
        r = setup_row + 2 + i
        set_cell(ws, f"B{r}", label, font=FONT_BODY_BOLD, border=BORDER_ALL, align=ALIGN_LEFT)
        cell = set_cell(ws, f"C{r}", default, font=FONT_BODY, fill=FILL_INPUT, border=BORDER_ALL, align=ALIGN_LEFT)
        ws.merge_cells(f"C{r}:D{r}")
        if kind == "date":
            cell.number_format = "yyyy-mm-dd"
        elif kind == "currency":
            cell.number_format = "$#,##0.00"
        value_cells[label] = f"C{r}"
        ws.row_dimensions[r].height = 22

    # Yes/No validation on Period Locked
    yn_dv = DataValidation(type="list", formula1=f'"{",".join(YES_NO)}"', allow_blank=False)
    yn_dv.add(value_cells["Period Locked in GL?"])
    ws.add_data_validation(yn_dv)

    # Defined names — these are what the other sheets reference
    add_defined_name(wb, "CompanyName", f"'README & Setup'!${value_cells['Company Name'][0]}${value_cells['Company Name'][1:]}")
    add_defined_name(wb, "CurrentPeriod", f"'README & Setup'!${value_cells['Current Period (e.g., May 2026)'][0]}${value_cells['Current Period (e.g., May 2026)'][1:]}")
    add_defined_name(wb, "PeriodEndDate", f"'README & Setup'!${value_cells['Period End Date'][0]}${value_cells['Period End Date'][1:]}")
    add_defined_name(wb, "Materiality", f"'README & Setup'!${value_cells['Performance Materiality ($)'][0]}${value_cells['Performance Materiality ($)'][1:]}")
    add_defined_name(wb, "PeriodLocked", f"'README & Setup'!${value_cells['Period Locked in GL?'][0]}${value_cells['Period Locked in GL?'][1:]}")

    # CTA / upsell
    cta_row = setup_row + 2 + len(setup_fields) + 2
    set_cell(ws, f"B{cta_row}", "Need more depth?", font=FONT_H2)
    upsells = [
        ("ASC 606 Commission Accrual Workbook", "50 deals, dynamic period selection, JE generator, rollforward — $79", "https://kdeskaccounting.gumroad.com/l/mwmwpe"),
        ("ASC 842 Lease Accounting Workbook", "20 leases, six-category JE generator, ROU rollforward — $97", "https://kdeskaccounting.gumroad.com/l/phxigq"),
        ("Fixed Asset Rollforward Workbook", "50 assets, four depreciation methods, audit-ready rollforward — $79", "https://kdeskaccounting.gumroad.com/l/fixed-asset-rollforward"),
        ("Runway Calculator (free, browser-based)", "ASC 606 + ASC 842 amortization estimates in seconds", "https://kdeskaccounting.com/calculator/"),
    ]
    for i, (name, desc, url) in enumerate(upsells):
        r = cta_row + 1 + i
        set_cell(ws, f"B{r}", name, font=FONT_BODY_BOLD, align=ALIGN_LEFT)
        cell = set_cell(ws, f"C{r}", desc, font=FONT_BODY, align=ALIGN_LEFT)
        ws.merge_cells(f"C{r}:E{r}")
        link_cell = set_cell(ws, f"F{r}", "Open →", font=FONT_LINK, align=ALIGN_CENTER)
        link_cell.hyperlink = url
        ws.row_dimensions[r].height = 22

    # Footer note
    footer_row = cta_row + 2 + len(upsells) + 1
    set_cell(ws, f"B{footer_row}", "Not legal, tax, or audit advice. Use within your firm's policies and your auditor's expectations.", font=FONT_NOTE, align=ALIGN_LEFT)
    ws.merge_cells(f"B{footer_row}:F{footer_row}")


def build_close_calendar(wb: Workbook) -> None:
    ws = wb.create_sheet("Close Calendar")
    set_col_widths(ws, {"A": 4, "B": 6, "C": 32, "D": 56, "E": 22, "F": 14, "G": 16, "H": 32})
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B5"

    # Title band
    ws.merge_cells("B2:H2")
    set_cell(ws, "B2", "Close Calendar", font=Font(name="Calibri", size=18, bold=True, color=COLOR_WHITE),
             fill=FILL_HEADER, align=Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:H3")
    set_cell(ws, "B3", '=CONCATENATE("Period: ",CurrentPeriod,"  |  Company: ",CompanyName)',
             font=FONT_NOTE, align=Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[3].height = 20

    # Header row at 5
    headers = ["#", "Phase", "Task", "Owner", "Due Date", "Status", "Notes"]
    write_table_header(ws, 5, headers, start_col=2)

    # Body rows
    start_row = 6
    for i, (phase, task, owner) in enumerate(TASKS):
        r = start_row + i
        set_cell(ws, f"B{r}", i + 1, font=FONT_BODY, align=ALIGN_CENTER, border=BORDER_ALL)
        set_cell(ws, f"C{r}", phase, font=FONT_BODY, align=ALIGN_LEFT_TOP, border=BORDER_ALL)
        set_cell(ws, f"D{r}", task, font=FONT_BODY, align=ALIGN_LEFT_TOP, border=BORDER_ALL)
        set_cell(ws, f"E{r}", owner, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_LEFT, border=BORDER_ALL)
        set_cell(ws, f"F{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_CENTER, border=BORDER_ALL, number_format="yyyy-mm-dd")
        set_cell(ws, f"G{r}", "Not Started", font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_CENTER, border=BORDER_ALL)
        set_cell(ws, f"H{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_LEFT_TOP, border=BORDER_ALL)
        ws.row_dimensions[r].height = 36

    end_row = start_row + len(TASKS) - 1

    # Status dropdown
    status_dv = DataValidation(type="list", formula1=f'"{",".join(STATUS_OPTIONS)}"', allow_blank=False)
    status_dv.add(f"G{start_row}:G{end_row}")
    ws.add_data_validation(status_dv)

    # Conditional formatting on Status
    status_range = f"G{start_row}:G{end_row}"
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Done"'], fill=PatternFill("solid", fgColor=COLOR_GREEN), font=Font(color=COLOR_GREEN_FG, bold=True)))
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"In Progress"'], fill=PatternFill("solid", fgColor=COLOR_YELLOW), font=Font(color=COLOR_YELLOW_FG, bold=True)))
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Blocked"'], fill=PatternFill("solid", fgColor=COLOR_RED), font=Font(color=COLOR_RED_FG, bold=True)))
    ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"N/A"'], fill=PatternFill("solid", fgColor=COLOR_GRAY), font=Font(color=COLOR_GRAY_FG, italic=True)))

    # Defined name for the status column — used by Sign-Off summary
    add_defined_name(wb, "TaskStatus", f"'Close Calendar'!$G${start_row}:$G${end_row}")
    add_defined_name(wb, "TaskList", f"'Close Calendar'!$D${start_row}:$D${end_row}")

    # Summary footer
    sum_row = end_row + 2
    set_cell(ws, f"C{sum_row}", "Totals", font=FONT_BODY_BOLD, align=ALIGN_RIGHT)
    set_cell(ws, f"D{sum_row}", f'=CONCATENATE(COUNTA(D{start_row}:D{end_row}), " tasks total  |  ", COUNTIF(G{start_row}:G{end_row},"Done"), " done  |  ", COUNTIF(G{start_row}:G{end_row},"In Progress"), " in progress  |  ", COUNTIF(G{start_row}:G{end_row},"Blocked"), " blocked  |  ", COUNTIF(G{start_row}:G{end_row},"N/A"), " N/A")',
             font=FONT_BODY_BOLD, align=ALIGN_LEFT, fill=FILL_SECTION, border=BORDER_ALL)
    ws.merge_cells(f"D{sum_row}:H{sum_row}")
    ws.row_dimensions[sum_row].height = 28


def build_reconciliations(wb: Workbook) -> None:
    ws = wb.create_sheet("Reconciliations")
    set_col_widths(ws, {"A": 4, "B": 6, "C": 34, "D": 34, "E": 18, "F": 18, "G": 16, "H": 16, "I": 16, "J": 18, "K": 18, "L": 32})
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B5"

    # Title band
    ws.merge_cells("B2:L2")
    set_cell(ws, "B2", "Reconciliations (Subledger → GL Tie-Out)", font=Font(name="Calibri", size=18, bold=True, color=COLOR_WHITE),
             fill=FILL_HEADER, align=Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:L3")
    set_cell(ws, "B3", '=CONCATENATE("Period: ",CurrentPeriod,"  |  Materiality threshold: $",TEXT(Materiality,"#,##0"))',
             font=FONT_NOTE, align=Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[3].height = 20

    # Headers
    headers = ["#", "Account", "Source / Schedule", "GL Balance", "Sub/Schedule Balance",
               "Variance", "Within Mat.?", "Status", "Preparer", "Reviewer", "Notes"]
    write_table_header(ws, 5, headers, start_col=2)

    start_row = 6
    for i, (account, source) in enumerate(RECONCILIATIONS):
        r = start_row + i
        set_cell(ws, f"B{r}", i + 1, font=FONT_BODY, align=ALIGN_CENTER, border=BORDER_ALL)
        set_cell(ws, f"C{r}", account, font=FONT_BODY, align=ALIGN_LEFT, border=BORDER_ALL)
        set_cell(ws, f"D{r}", source, font=FONT_BODY, align=ALIGN_LEFT, border=BORDER_ALL)
        # GL Balance (input)
        c = set_cell(ws, f"E{r}", 0, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_RIGHT, border=BORDER_ALL)
        c.number_format = "$#,##0.00;[Red]($#,##0.00);\"–\""
        # Sub/Schedule Balance (input)
        c = set_cell(ws, f"F{r}", 0, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_RIGHT, border=BORDER_ALL)
        c.number_format = "$#,##0.00;[Red]($#,##0.00);\"–\""
        # Variance (formula)
        c = set_cell(ws, f"G{r}", f"=E{r}-F{r}", font=FONT_BODY_BOLD, align=ALIGN_RIGHT, border=BORDER_ALL)
        c.number_format = "$#,##0.00;[Red]($#,##0.00);\"–\""
        # Within Materiality? (formula)
        set_cell(ws, f"H{r}", f'=IF(ABS(G{r})<=Materiality,"OK","FLAG")', font=FONT_BODY_BOLD, align=ALIGN_CENTER, border=BORDER_ALL)
        # Status dropdown
        set_cell(ws, f"I{r}", "Not Started", font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_CENTER, border=BORDER_ALL)
        # Preparer / Reviewer / Notes
        set_cell(ws, f"J{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_LEFT, border=BORDER_ALL)
        set_cell(ws, f"K{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_LEFT, border=BORDER_ALL)
        set_cell(ws, f"L{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_LEFT_TOP, border=BORDER_ALL)
        ws.row_dimensions[r].height = 26

    end_row = start_row + len(RECONCILIATIONS) - 1

    # Status dropdown
    status_dv = DataValidation(type="list", formula1=f'"{",".join(STATUS_OPTIONS)}"', allow_blank=False)
    status_dv.add(f"I{start_row}:I{end_row}")
    ws.add_data_validation(status_dv)

    # Conditional formatting on Within Materiality column
    flag_range = f"H{start_row}:H{end_row}"
    ws.conditional_formatting.add(flag_range, CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=COLOR_GREEN), font=Font(color=COLOR_GREEN_FG, bold=True)))
    ws.conditional_formatting.add(flag_range, CellIsRule(operator="equal", formula=['"FLAG"'], fill=PatternFill("solid", fgColor=COLOR_RED), font=Font(color=COLOR_RED_FG, bold=True)))

    # Status CF
    rec_status_range = f"I{start_row}:I{end_row}"
    ws.conditional_formatting.add(rec_status_range, CellIsRule(operator="equal", formula=['"Done"'], fill=PatternFill("solid", fgColor=COLOR_GREEN), font=Font(color=COLOR_GREEN_FG, bold=True)))
    ws.conditional_formatting.add(rec_status_range, CellIsRule(operator="equal", formula=['"In Progress"'], fill=PatternFill("solid", fgColor=COLOR_YELLOW), font=Font(color=COLOR_YELLOW_FG, bold=True)))
    ws.conditional_formatting.add(rec_status_range, CellIsRule(operator="equal", formula=['"Blocked"'], fill=PatternFill("solid", fgColor=COLOR_RED), font=Font(color=COLOR_RED_FG, bold=True)))

    # Defined name — used by Sign-Off
    add_defined_name(wb, "RecFlags", f"Reconciliations!$H${start_row}:$H${end_row}")
    add_defined_name(wb, "RecVariance", f"Reconciliations!$G${start_row}:$G${end_row}")
    add_defined_name(wb, "RecStatus", f"Reconciliations!$I${start_row}:$I${end_row}")

    # Summary footer
    sum_row = end_row + 2
    set_cell(ws, f"C{sum_row}", "Totals", font=FONT_BODY_BOLD, align=ALIGN_RIGHT)
    set_cell(ws, f"D{sum_row}", "GL Total", font=FONT_BODY_BOLD, align=ALIGN_RIGHT, fill=FILL_SECTION)
    c = set_cell(ws, f"E{sum_row}", f"=SUM(E{start_row}:E{end_row})", font=FONT_BODY_BOLD, align=ALIGN_RIGHT, fill=FILL_SECTION, border=BORDER_ALL)
    c.number_format = "$#,##0.00;[Red]($#,##0.00);\"–\""
    c = set_cell(ws, f"F{sum_row}", f"=SUM(F{start_row}:F{end_row})", font=FONT_BODY_BOLD, align=ALIGN_RIGHT, fill=FILL_SECTION, border=BORDER_ALL)
    c.number_format = "$#,##0.00;[Red]($#,##0.00);\"–\""
    c = set_cell(ws, f"G{sum_row}", f"=E{sum_row}-F{sum_row}", font=FONT_BODY_BOLD, align=ALIGN_RIGHT, fill=FILL_SECTION, border=BORDER_ALL)
    c.number_format = "$#,##0.00;[Red]($#,##0.00);\"–\""
    set_cell(ws, f"H{sum_row}", f'=CONCATENATE(COUNTIF(H{start_row}:H{end_row},"FLAG")," flags")', font=FONT_BODY_BOLD, align=ALIGN_CENTER, fill=FILL_SECTION, border=BORDER_ALL)
    ws.row_dimensions[sum_row].height = 28


def build_je_tracker(wb: Workbook) -> None:
    ws = wb.create_sheet("JE Tracker")
    set_col_widths(ws, {"A": 4, "B": 6, "C": 14, "D": 12, "E": 38, "F": 22, "G": 16, "H": 16, "I": 24, "J": 18, "K": 18, "L": 16})
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B5"

    ws.merge_cells("B2:L2")
    set_cell(ws, "B2", "Journal Entry Tracker", font=Font(name="Calibri", size=18, bold=True, color=COLOR_WHITE),
             fill=FILL_HEADER, align=Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:L3")
    set_cell(ws, "B3", '=CONCATENATE("Period: ",CurrentPeriod,"  |  Log every JE posted this period. Sanity check at the bottom catches one-sided entries.")',
             font=FONT_NOTE, align=Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[3].height = 20

    headers = ["#", "Date", "JE #", "Description", "Account", "Debit", "Credit", "Source", "Preparer", "Reviewer", "Status"]
    write_table_header(ws, 5, headers, start_col=2)

    start_row = 6
    rows = 50
    end_row = start_row + rows - 1

    for r in range(start_row, end_row + 1):
        set_cell(ws, f"B{r}", r - start_row + 1, font=FONT_BODY, align=ALIGN_CENTER, border=BORDER_ALL)
        c = set_cell(ws, f"C{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_CENTER, border=BORDER_ALL)
        c.number_format = "yyyy-mm-dd"
        set_cell(ws, f"D{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_CENTER, border=BORDER_ALL)
        set_cell(ws, f"E{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_LEFT, border=BORDER_ALL)
        set_cell(ws, f"F{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_LEFT, border=BORDER_ALL)
        c = set_cell(ws, f"G{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_RIGHT, border=BORDER_ALL)
        c.number_format = "$#,##0.00;[Red]($#,##0.00);\"–\""
        c = set_cell(ws, f"H{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_RIGHT, border=BORDER_ALL)
        c.number_format = "$#,##0.00;[Red]($#,##0.00);\"–\""
        set_cell(ws, f"I{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_CENTER, border=BORDER_ALL)
        set_cell(ws, f"J{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_LEFT, border=BORDER_ALL)
        set_cell(ws, f"K{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_LEFT, border=BORDER_ALL)
        set_cell(ws, f"L{r}", "Pending", font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_CENTER, border=BORDER_ALL)
        ws.row_dimensions[r].height = 22

    # Source dropdown
    source_dv = DataValidation(type="list", formula1=f'"{",".join(JE_SOURCE_OPTIONS)}"', allow_blank=True)
    source_dv.add(f"I{start_row}:I{end_row}")
    ws.add_data_validation(source_dv)

    # Status dropdown (JE-specific: Pending, Posted, Reversed)
    je_status_dv = DataValidation(type="list", formula1='"Pending,Posted,Reversed,Voided"', allow_blank=True)
    je_status_dv.add(f"L{start_row}:L{end_row}")
    ws.add_data_validation(je_status_dv)

    # JE status CF
    je_status_range = f"L{start_row}:L{end_row}"
    ws.conditional_formatting.add(je_status_range, CellIsRule(operator="equal", formula=['"Posted"'], fill=PatternFill("solid", fgColor=COLOR_GREEN), font=Font(color=COLOR_GREEN_FG, bold=True)))
    ws.conditional_formatting.add(je_status_range, CellIsRule(operator="equal", formula=['"Pending"'], fill=PatternFill("solid", fgColor=COLOR_YELLOW), font=Font(color=COLOR_YELLOW_FG, bold=True)))
    ws.conditional_formatting.add(je_status_range, CellIsRule(operator="equal", formula=['"Reversed"'], fill=PatternFill("solid", fgColor=COLOR_GRAY), font=Font(color=COLOR_GRAY_FG, italic=True)))

    # Sanity check footer
    sum_row = end_row + 2
    set_cell(ws, f"E{sum_row}", "Sanity Check:", font=FONT_BODY_BOLD, align=ALIGN_RIGHT, fill=FILL_SECTION, border=BORDER_ALL)
    set_cell(ws, f"F{sum_row}", "Sum of Debits / Credits / Difference", font=FONT_NOTE, align=ALIGN_LEFT, fill=FILL_SECTION, border=BORDER_ALL)
    c = set_cell(ws, f"G{sum_row}", f"=SUM(G{start_row}:G{end_row})", font=FONT_BODY_BOLD, align=ALIGN_RIGHT, fill=FILL_SECTION, border=BORDER_ALL)
    c.number_format = "$#,##0.00;[Red]($#,##0.00);\"–\""
    c = set_cell(ws, f"H{sum_row}", f"=SUM(H{start_row}:H{end_row})", font=FONT_BODY_BOLD, align=ALIGN_RIGHT, fill=FILL_SECTION, border=BORDER_ALL)
    c.number_format = "$#,##0.00;[Red]($#,##0.00);\"–\""
    c = set_cell(ws, f"I{sum_row}", f"=G{sum_row}-H{sum_row}", font=FONT_BODY_BOLD, align=ALIGN_RIGHT, fill=FILL_SECTION, border=BORDER_ALL)
    c.number_format = "$#,##0.00;[Red]($#,##0.00);\"–\""

    flag_row = sum_row + 1
    set_cell(ws, f"E{flag_row}", "Balanced?", font=FONT_BODY_BOLD, align=ALIGN_RIGHT)
    c = set_cell(ws, f"F{flag_row}", f'=IF(ROUND(I{sum_row},2)=0,"✓ Debits = Credits","✗ OUT OF BALANCE")', font=FONT_BODY_BOLD, align=ALIGN_LEFT, border=BORDER_ALL)
    # Conditional format on the balance cell
    bal_cell = f"F{flag_row}"
    ws.conditional_formatting.add(bal_cell, FormulaRule(formula=[f'ROUND(I{sum_row},2)=0'], fill=PatternFill("solid", fgColor=COLOR_GREEN), font=Font(color=COLOR_GREEN_FG, bold=True)))
    ws.conditional_formatting.add(bal_cell, FormulaRule(formula=[f'ROUND(I{sum_row},2)<>0'], fill=PatternFill("solid", fgColor=COLOR_RED), font=Font(color=COLOR_RED_FG, bold=True)))
    ws.row_dimensions[sum_row].height = 28
    ws.row_dimensions[flag_row].height = 24

    # Defined names for Sign-Off
    add_defined_name(wb, "JEStatus", f"'JE Tracker'!$L${start_row}:$L${end_row}")
    add_defined_name(wb, "JEDebitTotal", f"'JE Tracker'!$G${sum_row}")
    add_defined_name(wb, "JECreditTotal", f"'JE Tracker'!$H${sum_row}")
    add_defined_name(wb, "JEDifference", f"'JE Tracker'!$I${sum_row}")


def build_signoff(wb: Workbook) -> None:
    ws = wb.create_sheet("Sign-Off")
    set_col_widths(ws, {"A": 4, "B": 30, "C": 26, "D": 26, "E": 26, "F": 26})
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True

    # Title band
    ws.merge_cells("B2:F2")
    set_cell(ws, "B2", "Month-End Close Certification", font=Font(name="Calibri", size=20, bold=True, color=COLOR_WHITE),
             fill=FILL_HEADER, align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[2].height = 44

    ws.merge_cells("B3:F3")
    set_cell(ws, "B3", '=CONCATENATE(CompanyName,"  |  Period: ",CurrentPeriod,"  |  Period End: ",TEXT(PeriodEndDate,"yyyy-mm-dd"))',
             font=FONT_H2, align=ALIGN_CENTER)
    ws.row_dimensions[3].height = 28

    # Close metrics section
    set_cell(ws, "B5", "Close Metrics", font=FONT_H2)
    metrics = [
        ("Tasks total",            "=COUNTA(TaskList)"),
        ("Tasks completed",        '=COUNTIF(TaskStatus,"Done")'),
        ("Tasks in progress",      '=COUNTIF(TaskStatus,"In Progress")'),
        ("Tasks blocked",          '=COUNTIF(TaskStatus,"Blocked")'),
        ("Tasks N/A",              '=COUNTIF(TaskStatus,"N/A")'),
        ("Reconciliation flags",   '=COUNTIF(RecFlags,"FLAG")'),
        ("Reconciliations done",   '=COUNTIF(RecStatus,"Done")'),
        ("JE entries posted",      '=COUNTIF(JEStatus,"Posted")'),
        ("JE debit = credit?",     '=IF(ROUND(JEDifference,2)=0,"Yes","No — investigate")'),
    ]
    for i, (label, formula) in enumerate(metrics):
        r = 6 + i
        set_cell(ws, f"B{r}", label, font=FONT_BODY_BOLD, align=ALIGN_LEFT, border=BORDER_ALL)
        c = set_cell(ws, f"C{r}", formula, font=FONT_BODY_BOLD, align=ALIGN_RIGHT, fill=FILL_SECTION, border=BORDER_ALL)
        ws.merge_cells(f"C{r}:D{r}")
        ws.row_dimensions[r].height = 22

    # Close gate
    gate_row = 6 + len(metrics) + 1
    set_cell(ws, f"B{gate_row}", "Period Locked in GL?", font=FONT_BODY_BOLD, align=ALIGN_LEFT, border=BORDER_ALL)
    set_cell(ws, f"C{gate_row}", "=PeriodLocked", font=FONT_BODY_BOLD, align=ALIGN_CENTER, fill=FILL_SECTION, border=BORDER_ALL)
    ws.merge_cells(f"C{gate_row}:D{gate_row}")

    fs_row = gate_row + 1
    set_cell(ws, f"B{fs_row}", "Financial statements attached?", font=FONT_BODY_BOLD, align=ALIGN_LEFT, border=BORDER_ALL)
    set_cell(ws, f"C{fs_row}", "No", font=FONT_BODY_BOLD, align=ALIGN_CENTER, fill=FILL_INPUT, border=BORDER_ALL)
    ws.merge_cells(f"C{fs_row}:D{fs_row}")

    pkg_row = fs_row + 1
    set_cell(ws, f"B{pkg_row}", "Close package distributed?", font=FONT_BODY_BOLD, align=ALIGN_LEFT, border=BORDER_ALL)
    set_cell(ws, f"C{pkg_row}", "No", font=FONT_BODY_BOLD, align=ALIGN_CENTER, fill=FILL_INPUT, border=BORDER_ALL)
    ws.merge_cells(f"C{pkg_row}:D{pkg_row}")

    yn_dv = DataValidation(type="list", formula1=f'"{",".join(YES_NO)}"', allow_blank=False)
    yn_dv.add(f"C{fs_row}")
    yn_dv.add(f"C{pkg_row}")
    ws.add_data_validation(yn_dv)

    # Exception list
    excp_row = pkg_row + 2
    set_cell(ws, f"B{excp_row}", "Exception List (unresolved items, post-close follow-ups)", font=FONT_H2)
    for i in range(6):
        r = excp_row + 1 + i
        set_cell(ws, f"B{r}", i + 1, font=FONT_BODY, align=ALIGN_CENTER, border=BORDER_ALL)
        set_cell(ws, f"C{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_LEFT, border=BORDER_ALL)
        ws.merge_cells(f"C{r}:F{r}")
        ws.row_dimensions[r].height = 22

    # Signatures
    sig_row = excp_row + 1 + 6 + 2
    set_cell(ws, f"B{sig_row}", "Signatures", font=FONT_H2)

    sig_header_row = sig_row + 1
    write_table_header(ws, sig_header_row, ["Role", "Name", "Date", "Signature"], start_col=2)

    sig_roles = ["Preparer", "Reviewer", "Controller / CFO"]
    for i, role in enumerate(sig_roles):
        r = sig_header_row + 1 + i
        set_cell(ws, f"B{r}", role, font=FONT_BODY_BOLD, align=ALIGN_LEFT, border=BORDER_ALL)
        set_cell(ws, f"C{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_LEFT, border=BORDER_ALL)
        c = set_cell(ws, f"D{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_CENTER, border=BORDER_ALL)
        c.number_format = "yyyy-mm-dd"
        set_cell(ws, f"E{r}", None, font=FONT_BODY, fill=FILL_INPUT, align=ALIGN_LEFT, border=BORDER_ALL)
        ws.merge_cells(f"E{r}:F{r}")
        ws.row_dimensions[r].height = 32

    # Footer
    footer_row = sig_header_row + 1 + len(sig_roles) + 2
    set_cell(ws, f"B{footer_row}",
             "Not legal, tax, or audit advice. Use within your firm's policies and your auditor's expectations.  •  kdeskaccounting.com",
             font=FONT_NOTE, align=ALIGN_CENTER)
    ws.merge_cells(f"B{footer_row}:F{footer_row}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build_workbook(out_path: pathlib.Path) -> None:
    wb = Workbook()
    build_readme_setup(wb)
    build_close_calendar(wb)
    build_reconciliations(wb)
    build_je_tracker(wb)
    build_signoff(wb)

    # Set active sheet to README on open
    wb.active = 0

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", default=str(DIST / "month-end-close-checklist-v1.xlsx"))
    args = ap.parse_args()

    out = pathlib.Path(args.output)
    build_workbook(out)
    print(f"Wrote {out}  ({out.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
